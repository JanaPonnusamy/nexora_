"""Chunk 12 — Excel Export.

Builds the 6-sheet workbook exactly as frozen in
docs/Document_Extraction_Excel_Contract.md — column order/labels/types there
are binding; this module must not invent its own layout. Pure: takes
already-fetched rows (repository.get_imports_by_ids /
list_items_by_import_ids), returns bytes. No DB access, no file I/O — the
caller (service.py) owns persisting the result via storage.export_path().

`csv` format is Sheet 2 (Products) only, per the contract's own constraint
(a single table can't represent six sheets).
"""

import csv
import io
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_DATE_FMT = "%Y-%m-%d"
_DATETIME_FMT = "%Y-%m-%d %H:%M:%S"


def _slab_get(slab: dict, *keys):
    """gst_slab_breakup_json's exact key names aren't written by any chunk
    yet (Chunk 14/Integration is first) — tolerate either the Excel
    contract doc's short names (rate/taxable/cgst/...) or the GSTSummary
    model's names (gst_percent/taxable_amount/...) so this module works
    whichever one Integration ends up emitting."""
    for key in keys:
        if key in slab and slab[key] is not None:
            return slab[key]
    return None


def _as_date(value):
    if value is None:
        return None
    if isinstance(value, (date, datetime)):
        return value.strftime(_DATE_FMT)
    return str(value)[:10]


def _as_datetime(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime(_DATETIME_FMT)
    return str(value)


def _write_header_row(ws, headers):
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def _autosize(ws, headers):
    for col_idx, header in enumerate(headers, start=1):
        values = (
            str(ws.cell(row=row_idx, column=col_idx).value or "")
            for row_idx in range(2, min(ws.max_row, 200) + 1)
        )
        longest = max([len(header), *(len(value) for value in values)])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(36, max(12, longest + 2))
    ws.auto_filter.ref = ws.dimensions


def _format_pharmacy_products(ws):
    if ws.max_row < 2:
        return
    by_name = {cell.value: cell.column for cell in ws[1]}
    for name in ("Qty", "Free", "Mixed case(qty)"):
        for cell in ws.iter_cols(
            min_col=by_name[name], max_col=by_name[name], min_row=2, max_row=ws.max_row,
        ):
            for value in cell:
                value.number_format = "0.###"
    for name in (
        "Rate", "PDisc.", "Amount", "PDisc.Amount", "CSt.Amount", "St.Amount",
        "Goods Value", "SDisAmt", "CashDiscAmt", "Mrp", "PTR", "Net Rate",
        "Net Rate Amount", "OFFER RATE", "MRP VALUE", "PTS", "Sch.Dis",
        "Sch.DisAmt", "SplDis", "SplDisAmt", "Pdisc%", "SGST", "CGST", "IGST",
        "GST Cess", "Abate Perc", "Rate pack", "Calamity CESS", "Calamity %",
        "Extra Cess / Base Unit", "Extra Cess Amt",
    ):
        for cell in ws.iter_cols(
            min_col=by_name[name], max_col=by_name[name], min_row=2, max_row=ws.max_row,
        ):
            for value in cell:
                value.number_format = "0.00##"


# --------------------------------------------------------------------------
# Sheet 1 — Invoice Header
# --------------------------------------------------------------------------

_HEADER_COLUMNS = [
    "Import ID", "Supplier Name", "GST Number", "DL Number", "Invoice Number",
    "Invoice Date", "Invoice Type", "Order Number", "Transport", "Salesman",
    "Credit Days", "Gross Amount", "Discount Amount", "Scheme Discount",
    "Cash Discount", "Taxable Amount", "CGST Amount", "SGST Amount",
    "IGST Amount", "CESS Amount", "Round Off", "Net Amount", "Item Count",
    "Total Quantity", "IRN Number", "Ack Number", "Ack Date", "Validation Status",
]


def _header_row(doc: dict) -> list:
    return [
        doc["import_id"], doc.get("supplier_name"), doc.get("gst_number"), doc.get("dl_number"),
        doc.get("invoice_number"), _as_date(doc.get("invoice_date")), doc.get("invoice_type"),
        doc.get("order_number"), doc.get("transport"), doc.get("salesman"), doc.get("credit_days"),
        doc.get("gross_amount"), doc.get("discount_amount"), doc.get("scheme_discount"),
        doc.get("cash_discount"), doc.get("taxable_amount"), doc.get("cgst_amount"),
        doc.get("sgst_amount"), doc.get("igst_amount"), doc.get("cess_amount"), doc.get("round_off"),
        doc.get("net_amount"), doc.get("item_count"), doc.get("total_quantity"),
        doc.get("irn_number"), doc.get("ack_number"), _as_date(doc.get("ack_date")),
        doc.get("validation_status"),
    ]


# --------------------------------------------------------------------------
# Sheet 2 — Products (pharmacy import layout)
# --------------------------------------------------------------------------

# Column labels and order supplied by the pharmacy import screen. Keep this
# as an explicit compatibility contract: downstream desktop software commonly
# maps by ordinal position, not by a resilient column-name lookup.
_PRODUCT_COLUMNS = [
    "No", "Product Name", "Packing", "Batch", "Qty", "Free", "Rate", "Tax %",
    "PDisc.", "Amount", "Mfr", "DC", "PDisc.Amount", "CSt.Amount", "St.Amount",
    "Goods Value", "Prod. Code", "SDisAmt", "CashDiscAmt", "Exp.dt", "Mrp",
    "Sales", "Drugs", "PTR", "Batchid", "Net Rate", "Net Rate Amount", "OFFER",
    "OFFER RATE", "MRP VALUE", "PackCode", "PTS", "DC-REF", "OfferType", "Cdis",
    "Sdis", "Case", "GodownID", "Repl", "SchOfferId", "OfferSlno", "Sch.Dis",
    "Sch.DisAmt", "MixedCase", "Parentid", "autodisc", "volume", "SplOfferid",
    "SplDis", "SplDisAmt", "Pdisc%", "TotalVolume", "SO no", "SO Sno",
    "Gross Weight", "Edited-Spldisc", "Edited-Schdisc", "SpL Zero", "Sch Zero",
    "SplZeroId", "SchZeroID", "SGST", "CGST", "IGST", "GST Cess", "Abate Perc",
    "GST Based on", "HSN Code", "Rate Changable", "Mixed case(qty)", "Rate pack",
    "Alias Code", "Calamity CESS", "Remarks", "Calamity %",
    "Extra Cess / Base Unit", "Extra Cess Amt", "Tax Slab Code", "HQ Approval Status",
]


# The original normalized OCR view is retained in a separate sheet. It keeps
# Import ID + Invoice Number on every row for batch-export traceability while
# Products stays exactly compatible with the requested import layout, which
# has no invoice-number column.
_OCR_PRODUCT_COLUMNS = [
    "Import ID", "Invoice Number", "Line No", "Product Code", "Product Name", "Pack",
    "HSN Code", "Batch Number", "Expiry Date", "Quantity", "Free Quantity", "PTR",
    "Purchase Rate", "MRP", "GST %", "Discount %", "Discount Amount", "Amount", "Confidence",
]


def _ocr_product_row(item: dict, invoice_number) -> list:
    return [
        item["import_id"], invoice_number, item["line_number"], item.get("product_code"),
        item.get("normalized_product_name") or item.get("ocr_product_name"), item.get("pack"),
        item.get("hsn_code"), item.get("batch_number"),
        _as_date(item.get("expiry_date")) or item.get("expiry_raw"),
        item.get("quantity"), item.get("free_quantity"), item.get("ptr"), item.get("purchase_rate"),
        item.get("mrp"), item.get("gst_percent"), item.get("discount_percent"),
        item.get("discount_amount"), item.get("amount"), item.get("confidence"),
    ]


def _multiply(left, right):
    if left is None or right is None:
        return None
    return round(left * right, 4)


def _purchase_rate_for_export(item: dict):
    """Return the best reviewed unit cost for the pharmacy Rate column."""
    if item.get("purchase_rate") is not None:
        return item["purchase_rate"]
    amount = item.get("amount")
    quantity = item.get("quantity")
    gst_percent = item.get("gst_percent")
    if amount is not None and quantity not in (None, 0) and gst_percent is not None:
        # In the supplied invoice `Total` includes GST while `GST Base` is the
        # taxable line value. Recover that base rate when OCR has PTR but no
        # separate purchase-rate column.
        return round(amount / (1 + gst_percent / 100) / quantity, 4)
    return item.get("ptr")


def _discount_amount_for_export(item: dict, rate):
    if item.get("discount_amount") is not None:
        return item["discount_amount"]
    quantity = item.get("quantity")
    percent = item.get("discount_percent")
    if rate is None or quantity is None or percent is None:
        return None
    return round(rate * quantity * percent / 100, 2)


def _expiry_for_pharmacy(item: dict):
    """The target format shows pharma expiry as MM/YY, not a day-level date."""
    value = item.get("expiry_date")
    if isinstance(value, (date, datetime)):
        return value.strftime("%m/%y")
    if value:
        text = str(value)
        try:
            parsed = datetime.strptime(text[:10], _DATE_FMT)
            return parsed.strftime("%m/%y")
        except ValueError:
            pass
    return item.get("expiry_raw")


def _line_tax(doc: dict, item: dict):
    """Best-effort line GST using reviewed values, never unverified guesses."""
    rate = _purchase_rate_for_export(item)
    qty = item.get("quantity")
    gst_percent = item.get("gst_percent")
    if rate is None or qty is None or gst_percent is None:
        return None, None, None
    taxable = rate * qty - (_discount_amount_for_export(item, rate) or 0)
    tax = round(taxable * gst_percent / 100, 2)
    if doc.get("igst_amount") not in (None, 0):
        return None, None, tax
    if doc.get("cgst_amount") not in (None, 0) or doc.get("sgst_amount") not in (None, 0):
        # Indian intra-state GST components are rounded independently.
        half = round(taxable * gst_percent / 200, 2)
        return half, half, None
    return None, None, None


def _product_row(item: dict, doc: dict) -> list:
    rate = _purchase_rate_for_export(item)
    qty = item.get("quantity")
    mrp = item.get("mrp")
    sgst, cgst, igst = _line_tax(doc, item)
    gst_percent = item.get("gst_percent")
    discount_amount = _discount_amount_for_export(item, rate)
    goods_value = _multiply(rate, qty)
    line_tax = sum(value or 0 for value in (sgst, cgst, igst))
    amount = item.get("amount")
    net_rate_amount = (
        round(amount + (discount_amount or 0), 2)
        if amount is not None
        else (round(goods_value + line_tax, 2) if goods_value is not None else None)
    )
    net_rate = (
        round(net_rate_amount / qty, 4)
        if net_rate_amount is not None and qty not in (None, 0) else None
    )

    # Defaults mirror the target system's exports: flags are booleans, known
    # numeric-but-unavailable fields are zero, and identifiers remain blank.
    values = {column: None for column in _PRODUCT_COLUMNS}
    values.update({
        "No": item["line_number"],
        "Product Name": item.get("normalized_product_name") or item.get("ocr_product_name"),
        "Packing": item.get("pack"),
        "Batch": item.get("batch_number"),
        "Qty": qty,
        "Free": item.get("free_quantity"),
        "Rate": rate,
        "Tax %": gst_percent,
        "PDisc.": item.get("discount_percent"),
        "Amount": item.get("amount"),
        "DC": False,
        "PDisc.Amount": discount_amount,
        "Goods Value": goods_value,
        "Prod. Code": item.get("product_code"),
        "Exp.dt": _expiry_for_pharmacy(item),
        "Mrp": mrp,
        "Sales": "GST" if gst_percent is not None else None,
        "PTR": item.get("ptr"),
        "Net Rate": net_rate,
        "Net Rate Amount": net_rate_amount,
        "MRP VALUE": _multiply(mrp, qty),
        "Repl": False,
        "SGST": sgst or 0,
        "CGST": cgst or 0,
        "IGST": igst or 0,
        "HSN Code": item.get("hsn_code"),
        "Rate Changable": False,
        "Tax Slab Code": 0,
    })
    for column in (
        "CSt.Amount", "St.Amount", "SDisAmt", "CashDiscAmt",
        "Cdis", "Sdis", "Case", "GodownID",
        "SchOfferId", "OfferSlno", "Sch.Dis", "Sch.DisAmt", "MixedCase",
        "Parentid", "autodisc", "volume", "SplOfferid", "SplDis",
        "SplDisAmt", "Pdisc%", "TotalVolume", "Gross Weight",
        "Edited-Spldisc", "Edited-Schdisc", "SpL Zero", "Sch Zero",
        "SplZeroId", "SchZeroID", "GST Cess", "Abate Perc",
        "Mixed case(qty)", "Rate pack", "Calamity CESS", "Calamity %",
        "Extra Cess / Base Unit", "Extra Cess Amt",
    ):
        values[column] = 0
    return [values[column] for column in _PRODUCT_COLUMNS]


# --------------------------------------------------------------------------
# Sheet 4 — GST Summary
# --------------------------------------------------------------------------

_GST_SUMMARY_COLUMNS = [
    "Import ID", "Invoice Number", "GST Rate %", "Taxable Amount", "CGST Amount",
    "SGST Amount", "IGST Amount", "Total GST Amount", "Net Amount",
]


def _gst_summary_rows(doc: dict) -> list:
    slabs = doc.get("gst_slab_breakup_json") or []
    rows = []
    for slab in slabs:
        cgst = _slab_get(slab, "cgst", "cgst_amount") or 0
        sgst = _slab_get(slab, "sgst", "sgst_amount") or 0
        igst = _slab_get(slab, "igst", "igst_amount") or 0
        total = _slab_get(slab, "total", "total_amount")
        if total is None:
            total = cgst + sgst + igst
        rows.append([
            doc["import_id"], doc.get("invoice_number"),
            _slab_get(slab, "rate", "gst_percent"), _slab_get(slab, "taxable", "taxable_amount"),
            _slab_get(slab, "cgst", "cgst_amount"), _slab_get(slab, "sgst", "sgst_amount"),
            _slab_get(slab, "igst", "igst_amount"), total, _slab_get(slab, "net", "net_amount"),
        ])
    return rows


# --------------------------------------------------------------------------
# Sheet 5 — Validation Errors
# --------------------------------------------------------------------------

_VALIDATION_COLUMNS = [
    "Import ID", "Invoice Number", "Rule Code", "Severity", "Field", "Line No",
    "Message", "Expected Value", "Actual Value",
]


def _validation_rows(doc: dict, line_by_item_id: dict) -> list:
    findings = doc.get("validation_json") or []
    rows = []
    for finding in findings:
        item_id = finding.get("item_id")
        rows.append([
            doc["import_id"], doc.get("invoice_number"), finding.get("rule_code"),
            finding.get("severity"), finding.get("field"),
            line_by_item_id.get(item_id) if item_id is not None else None,
            finding.get("message"), finding.get("expected_value"), finding.get("actual_value"),
        ])
    return rows


# --------------------------------------------------------------------------
# Sheet 6 — OCR Metadata
# --------------------------------------------------------------------------

_OCR_METADATA_COLUMNS = [
    "Import ID", "Invoice Number", "Source Type", "Page Count", "OCR Confidence %",
    "Supplier Match Method", "Supplier Match Confidence %", "Uploaded At", "Reviewed At", "Exported At",
]


def _ocr_metadata_row(doc: dict, exported_at: datetime) -> list:
    return [
        doc["import_id"], doc.get("invoice_number"), doc.get("source_type"), doc.get("page_count"),
        doc.get("ocr_confidence"), doc.get("supplier_match_method"), doc.get("supplier_match_confidence"),
        _as_datetime(doc.get("uploaded_at")), _as_datetime(doc.get("reviewed_at")),
        exported_at.strftime(_DATETIME_FMT),
    ]


# --------------------------------------------------------------------------
# Public builders
# --------------------------------------------------------------------------

def build_workbook(imports: list, items: list, exported_at: datetime) -> bytes:
    """`imports`/`items` are raw repository rows (get_imports_by_ids /
    list_items_by_import_ids — the latter WITHOUT excluded rows filtered out,
    since Sheet 4 findings may reference an item that was excluded after the
    finding was raised); this function filters Sheet 2 to non-excluded rows
    itself."""
    imports_by_id = {doc["import_id"]: doc for doc in imports}
    line_by_item_id = {item["item_id"]: item["line_number"] for item in items}

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Invoice Header"
    _write_header_row(ws1, _HEADER_COLUMNS)
    for doc in imports:
        ws1.append(_header_row(doc))
    _autosize(ws1, _HEADER_COLUMNS)

    ws2 = wb.create_sheet("Products")
    _write_header_row(ws2, _PRODUCT_COLUMNS)
    for item in items:
        if item.get("is_excluded"):
            continue
        doc = imports_by_id.get(item["import_id"], {})
        ws2.append(_product_row(item, doc))
    _autosize(ws2, _PRODUCT_COLUMNS)
    _format_pharmacy_products(ws2)

    ws2b = wb.create_sheet("OCR Products")
    _write_header_row(ws2b, _OCR_PRODUCT_COLUMNS)
    for item in items:
        if item.get("is_excluded"):
            continue
        doc = imports_by_id.get(item["import_id"], {})
        ws2b.append(_ocr_product_row(item, doc.get("invoice_number")))
    _autosize(ws2b, _OCR_PRODUCT_COLUMNS)

    ws3 = wb.create_sheet("GST Summary")
    _write_header_row(ws3, _GST_SUMMARY_COLUMNS)
    for doc in imports:
        for row in _gst_summary_rows(doc):
            ws3.append(row)
    _autosize(ws3, _GST_SUMMARY_COLUMNS)

    ws4 = wb.create_sheet("Validation")
    _write_header_row(ws4, _VALIDATION_COLUMNS)
    for doc in imports:
        for row in _validation_rows(doc, line_by_item_id):
            ws4.append(row)
    _autosize(ws4, _VALIDATION_COLUMNS)

    ws5 = wb.create_sheet("OCR Metadata")
    _write_header_row(ws5, _OCR_METADATA_COLUMNS)
    for doc in imports:
        ws5.append(_ocr_metadata_row(doc, exported_at))
    _autosize(ws5, _OCR_METADATA_COLUMNS)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_products_csv(imports: list, items: list) -> bytes:
    imports_by_id = {doc["import_id"]: doc for doc in imports}
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(_PRODUCT_COLUMNS)
    for item in items:
        if item.get("is_excluded"):
            continue
        doc = imports_by_id.get(item["import_id"], {})
        writer.writerow(_product_row(item, doc))
    return buffer.getvalue().encode("utf-8-sig")  # BOM so Excel opens UTF-8 CSV correctly
