"""Bulk-assign + Excel export for the Order Workspace "By Supplier" screen.

Ports the tail of the legacy WinForms Form1.btnExport_Click sequence:
UpdateDatabase() (bulk assign every grid row to the current supplier) then
ExportSelectedColumnsFromGrid() (the Excel build) -- Form1.vb, OrderManagement
WinForms project, ~lines 4441-4713.

Deliberately NOT ported: UpdateSupplierStockAfterExport(), which decremented
SupplierStock.stock by OrderQty*Pack after every export. Nexora already has an
independent, sync-driven supplier_stock pipeline (see the Supplier Live Stock
module); reproducing the legacy manual decrement here risks double-counting
against that sync, so this module intentionally stops once the Excel file is
built. This is an explicit product decision, not an oversight.
"""
import datetime
import io
import logging
import zipfile

from modules.legacy_order import repository

logger = logging.getLogger(__name__)

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
ZIP_MEDIA_TYPE = "application/zip"


class NoExportableRows(Exception):
    """No OrderQty>0 rows for this store/supplier -- nothing to assign or export."""


class ExportGenerationFailed(Exception):
    """Bulk assignment already committed; only the Excel build failed."""


class ExportResult:
    def __init__(self, content, filename, media_type, count, supplier_name):
        self.content = content
        self.filename = filename
        self.media_type = media_type
        self.count = count
        self.supplier_name = supplier_name


# --------------------------------------------------------------------------
# Category prediction -- port of Form1.PredictExportCategory /
# NormalizeExportCategory (Form1.vb ~lines 4407-4435).
# --------------------------------------------------------------------------

def _normalize_category(raw):
    if not raw or not raw.strip():
        return ""
    t = raw.strip().upper()
    for needle, tag in (
        ("TAB", "TAB"), ("SYP", "SYP"), ("SYRUP", "SYP"), ("PASTE", "PASTE"),
        ("VET", "VET"), ("INH", "INH"), ("RC", "RC"), ("RESP", "RESP"),
    ):
        if needle in t:
            return tag
    return ""


def _predict_category(sub_location, product_name, unit_description):
    from_sub = _normalize_category(sub_location)
    if from_sub:
        return from_sub
    text = f"{product_name or ''} {unit_description or ''}".upper()
    if "ROTACAP" in text:
        return "RC"
    if "INHALER" in text or "INHALLER" in text or " INH" in text:
        return "INH"
    if "RESPULE" in text or "RESP " in text or text.endswith("RESP"):
        return "RESP"
    if "PASTE" in text:
        return "PASTE"
    if "SUSPENSION" in text or "SYRUP" in text or " SYP" in text or text.endswith("SYP"):
        return "SYP"
    if "VET" in text:
        return "VET"
    if "TABLET" in text or " TAB" in text or text.endswith("TAB"):
        return "TAB"
    return "OTHER"


# --------------------------------------------------------------------------
# Filename -- port of Form1.SafeFileName (Form1.vb ~lines 4656-4665).
# --------------------------------------------------------------------------

_INVALID_FILENAME_CHARS = set('<>:"/\\|?*') | {chr(c) for c in range(32)}


def _safe_file_name(name):
    cleaned = "".join("_" if ch in _INVALID_FILENAME_CHARS else ch for ch in name)
    cleaned = cleaned.replace("[", "(").replace("]", ")")
    return cleaned.strip()


# --------------------------------------------------------------------------
# Workbook -- port of Form1.ExportSelectedColumnsFromGrid's column layout,
# header styling and S.No./OrderQty cell fills (Form1.vb ~lines 4441-4634).
# ProductCode is additionally forced to a TEXT cell (NumberFormat "@") --
# the legacy routine never did this (ProductCode is a plain int column in
# this schema), kept here only as a cheap safety margin.
# --------------------------------------------------------------------------

def _format_product_code(value):
    """ordermanagement.productcode round-trips through pyodbc as a Python
    float (e.g. 9810.0), not an int, even though it's always a whole number.
    str(9810.0) == "9810.0" -- writing that straight into the text cell would
    silently corrupt every exported ProductCode with a bogus ".0" suffix.
    Only numeric types get this treatment -- a string value is left exactly
    as-is (never round-tripped through float()), so a genuine text code with
    leading zeros is never touched."""
    if value is None:
        return ""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if float(value) == int(value):
            return str(int(value))
    return str(value)


def _build_workbook(rows, has_rack):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    headers = ["S.No.", "ProductName", "OrderQty", "SaleUnit", "MRP", "ProductCode", "Discount", "Category", "SubLocation"]
    if has_rack:
        headers.append("Rack")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    sno_fill = PatternFill("solid", fgColor="FFFFE0")
    qty_fill = PatternFill("solid", fgColor="90EE90")

    for i, row in enumerate(rows, start=1):
        values = [
            i,
            row.get("ProductName") or "",
            row.get("OrderQty") or 0,
            row.get("SaleUnit") or "",
            row.get("MRP") or 0,
            _format_product_code(row.get("ProductCode")),
            row.get("Discount") if has_rack else "",
            row.get("Category") or "",
            row.get("SubLocation") or "",
        ]
        if has_rack:
            values.append(row.get("Rack") or "")
        ws.append(values)

        excel_row = ws[i + 1]
        excel_row[0].fill = sno_fill
        excel_row[0].font = Font(bold=True)
        excel_row[0].alignment = Alignment(horizontal="center")
        excel_row[2].fill = qty_fill
        excel_row[2].font = Font(bold=True)
        excel_row[2].alignment = Alignment(horizontal="center")
        code_cell = excel_row[5]
        code_cell.number_format = "@"
        code_cell.value = _format_product_code(row.get("ProductCode"))

    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_order(store_name, supplier_code, supplier_name, mode, split_size, actor):
    """Resolve + validate -> bulk assign (commit) -> build Excel -> return.

    These are two separate steps, not one transaction: once the bulk assign
    commits, the assignment is real even if the Excel build fails afterward.
    A build failure raises ExportGenerationFailed with the assigned count so
    the caller can report it plainly rather than pretending nothing happened.
    """
    rows = repository.orders_by_supplier(store_name, supplier_code, mode)
    # orders_by_supplier's own SQL already restricts to status=0 AND
    # orderqty>0 (both modes) -- this second, explicit check doesn't change
    # today's behavior, it just stops this function from silently trusting
    # that upstream guarantee forever. If that SQL ever changes, an
    # already-assigned or non-positive row still can't reach the Excel file
    # or the bulk-assign product_codes list.
    rows = [r for r in rows if (r.get("OrderQty") or 0) > 0 and int(r.get("Status") or 0) == 0]
    if not rows:
        raise NoExportableRows("No products with a positive order quantity to export for this supplier.")

    has_rack = mode == "stock"
    for r in rows:
        r["Category"] = _predict_category(r.get("SubLocation"), r.get("ProductName"), r.get("UnitDescription"))

    if has_rack:
        rows.sort(key=lambda r: (0 if (r.get("Rack") or "").strip() else 1, (r.get("Rack") or "").strip(), r.get("ProductName") or ""))
    else:
        rows.sort(key=lambda r: r.get("ProductName") or "")

    product_codes = [r["ProductCode"] for r in rows]
    assigned_count = repository.bulk_assign_rows(store_name, supplier_code, supplier_name, product_codes, actor)

    date_part = datetime.date.today().strftime("%d%m%Y")
    base_name = _safe_file_name(f"{supplier_name} {store_name}   {date_part}")

    try:
        split_size = int(split_size or 0)
    except (TypeError, ValueError):
        split_size = 0

    if 0 < split_size < len(rows):
        chunks = [rows[i:i + split_size] for i in range(0, len(rows), split_size)]
    else:
        chunks = [rows]

    try:
        if len(chunks) == 1:
            content = _build_workbook(chunks[0], has_rack)
            filename = f"{base_name}.xlsx"
            media_type = XLSX_MEDIA_TYPE
        else:
            buf = io.BytesIO()
            with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for idx, chunk in enumerate(chunks, start=1):
                    part_name = f"{base_name} Part {idx} of {len(chunks)}.xlsx"
                    zf.writestr(part_name, _build_workbook(chunk, has_rack))
            content = buf.getvalue()
            filename = f"{base_name}.zip"
            media_type = ZIP_MEDIA_TYPE
    except Exception as exc:
        logger.exception(
            "Order export: bulk-assign committed (%s rows) but workbook build failed "
            "for store=%s supplier=%s", assigned_count, store_name, supplier_code,
        )
        raise ExportGenerationFailed(
            f"Bulk assignment succeeded ({assigned_count} products assigned), but Excel "
            "generation failed. The assignment is real -- check the Assigned tab, then retry the export."
        ) from exc

    return ExportResult(content, filename, media_type, assigned_count, supplier_name)
