"""Styled .xlsx generation for the Time Report module.

Ported from the standalone COSEC Flask app's excel_export.py. Consumes the
exact same dicts the JSON API returns, so the download never drifts from what
the screen shows.
"""
import io

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

from modules.time_report import config

BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FILL = PatternFill("solid", fgColor=config.COLOR_HEADER)
SUBHEAD_FILL = PatternFill("solid", fgColor=config.COLOR_SUBHEAD)
TOTAL_FILL = PatternFill("solid", fgColor=config.COLOR_TOTAL)

N_PUNCH = 8       # raw punch columns shown
NAME_MAX = 30


def _fill(hex_):
    return PatternFill("solid", fgColor=hex_)


def _white_font(hex_text):
    return Font(color=hex_text)


def _autosize(ws):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        width = 8
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell, MergedCell):
                continue
            if cell.value is not None:
                width = max(width, len(str(cell.value)) + 2)
        ws.column_dimensions[letter].width = min(width, 40)


def _write_legend(ws, start_row, legend):
    r = start_row + 1
    ws.cell(row=r, column=1, value="Legend:").font = BOLD
    r += 1
    for item in legend:
        c = ws.cell(row=r, column=1, value=f"{item['code']}  {item['label']}")
        c.fill = _fill(item["hex"])
        c.border = BORDER
        r += 1
    return r


def _save(wb) -> io.BytesIO:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============================================================
# 1. DAILY
# ============================================================
def daily_xlsx(data) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily"
    headers = ["ID", "Name"] + [str(i + 1) for i in range(N_PUNCH)] + ["Hrs"]
    ncol = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    t = ws.cell(row=1, column=1, value=data["period"])
    t.font = BOLD
    t.alignment = CENTER

    row = 2
    for dept in data["departments"]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
        c = ws.cell(row=row, column=1, value=dept["name"])
        c.fill = HEADER_FILL
        c.font = BOLD
        c.alignment = CENTER
        c.border = BORDER
        row += 1
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.fill = SUBHEAD_FILL
            c.font = BOLD
            c.alignment = CENTER
            c.border = BORDER
        row += 1
        for d in dept["rows"]:
            p = (d["punches"] + [""] * N_PUNCH)[:N_PUNCH]
            vals = [d["user_id"], d["name"][:NAME_MAX]] + p + [d["work_hm"]]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=v)
                c.border = BORDER
                c.alignment = LEFT if col == 2 else CENTER
                c.fill = _fill(d["status_hex"])
                c.font = _white_font(d["status_text"])
            row += 1
        tt = dept["totals"]
        m = ws.cell(row=row, column=1,
                    value=f"Present {tt['present']} | Absent {tt['absent']} | "
                          f"Miss Punch {tt['miss_punch']} | Late {tt['late']}")
        m.font = BOLD
        m.fill = _fill(config.COLOR_TOTAL)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
        row += 2

    _autosize(ws)
    ws.column_dimensions["B"].width = 26
    for i in range(3, 3 + N_PUNCH):
        ws.column_dimensions[get_column_letter(i)].width = 7
    _write_legend(ws, row, data["legend"])
    return _save(wb)


# ============================================================
# 2. MONTHLY MUSTER
# ============================================================
def monthly_xlsx(data) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly"
    days = data["days"]
    fixed = ["UserID", "Name", "Department"]
    tail = ["P", "Abs", "Miss", "Late", "Total"]
    ncol = len(fixed) + len(days) + len(tail)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    t = ws.cell(row=1, column=1, value=f"{data['title']}  -  {data['month_name']}")
    t.font = BOLD
    t.fill = HEADER_FILL
    t.alignment = CENTER

    r = 2
    col = 1
    for h in fixed:
        c = ws.cell(row=r, column=col, value=h)
        c.fill = SUBHEAD_FILL; c.font = BOLD; c.alignment = CENTER; c.border = BORDER
        col += 1
    for day in days:
        c = ws.cell(row=r, column=col, value=day)
        c.fill = SUBHEAD_FILL; c.font = BOLD; c.alignment = CENTER; c.border = BORDER
        col += 1
    for h in tail:
        c = ws.cell(row=r, column=col, value=h)
        c.fill = SUBHEAD_FILL; c.font = BOLD; c.alignment = CENTER; c.border = BORDER
        col += 1

    r = 3
    for u in data["rows"]:
        col = 1
        for v in (u["user_id"], u["name"], u["department"]):
            c = ws.cell(row=r, column=col, value=v)
            c.border = BORDER; c.alignment = LEFT if col == 2 else CENTER
            col += 1
        for cell in u["cells"]:
            c = ws.cell(row=r, column=col, value=cell["code"])
            c.border = BORDER; c.alignment = CENTER
            if cell["code"]:
                c.fill = _fill(cell["hex"])
                c.font = BOLD
            col += 1
        for v in (u["present"], u["absent"], u["miss_punch"], u["late"], u["total_hm"]):
            c = ws.cell(row=r, column=col, value=v)
            c.border = BORDER; c.alignment = CENTER; c.font = BOLD
            col += 1
        r += 1

    _autosize(ws)
    for i in range(len(fixed) + 1, len(fixed) + len(days) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 4
    _write_legend(ws, r, data["legend"])
    return _save(wb)


# ============================================================
# 3 & 4. flat tabular reports (miss punch, user detail/summary, inactive)
# ============================================================
def table_xlsx(title, headers, rows, legend=None) -> io.BytesIO:
    """Generic styled table. rows: list of (list_of_values, hex_or_None)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ncol = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    t = ws.cell(row=1, column=1, value=title)
    t.font = BOLD; t.fill = HEADER_FILL; t.alignment = CENTER

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = SUBHEAD_FILL; c.font = BOLD; c.alignment = CENTER; c.border = BORDER

    r = 3
    for values, hex_ in rows:
        for col, v in enumerate(values, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.border = BORDER
            c.alignment = CENTER if col != 2 else LEFT
            if hex_:
                c.fill = _fill(hex_)
        r += 1

    _autosize(ws)
    if legend:
        _write_legend(ws, r, legend)
    return _save(wb)
