"""Shelf Sorting & Excel Split.

Takes the existing export dataset for a whole Refresh, sorts it by shelf
category -> SubLocation -> ProductName (see export_document_service), and
splits it into consecutive .xlsx files of at most 16 products each — the exact
same Purchase Order layout the normal export produces, just reordered and cut
into pick-sized sheets. A single file is returned as-is; two or more are
bundled into one ZIP.

Reuses export_repository (to resolve the assignments) and
export_document_service (the one and only Excel engine) — no export logic is
duplicated here.
"""

import base64
import logging
import re
import zipfile
from copy import copy
from io import BytesIO

from fastapi import HTTPException

from config.database import get_connection
from modules.procurement import export_repository
from modules.procurement import export_document_service as docs
from modules.procurement import product_category_repository as pcat

logger = logging.getLogger("procurement.shelf_sort")

_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_ZIP = "application/zip"
_MAX_PER_FILE = 16

# Accepted header labels (normalised to lowercase alphanumerics) for the upload
# path — covers this app's own export plus common billing-software wording.
_ALIAS_CODE = ("productcode", "itemcode", "prodcode", "code", "sku")
_ALIAS_NAME = ("productname", "itemname", "productdescription", "itemdescription",
               "description", "particulars", "product", "item", "name")
_ALIAS_SUBLOC = ("sublocation", "subloc", "location", "rack", "shelf", "bin")
_ALIAS_UNIT = ("unitdescription", "unit", "uom", "packing", "pack", "unitdesc")
_ALIAS_SNO = ("sno", "slno", "srno", "serialno", "srlno", "slno")


def bundle_response(files, total, store_name):
    """([(name, bytes)], total) -> the (content, filename, media_type, total,
    file_count) tuple for a single download: a lone file streams directly,
    several are zipped."""
    if len(files) == 1:
        name, content = files[0]
        return content, name, _XLSX, total, 1
    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, content in files:
            zf.writestr(name, content)
    return buf.getvalue(), f"{docs._safe_name(store_name)}_Sorted.zip", _ZIP, total, len(files)


def json_payload(files, total):
    """The split files as a JSON-safe manifest (base64 content) so the desktop
    app can write each one INDIVIDUALLY into a chosen output folder (including
    a UNC/network path) rather than downloading a single zip."""
    return {
        "files": [{"name": name, "content_b64": base64.b64encode(content).decode("ascii")}
                  for name, content in files],
        "total_products": total,
        "file_count": len(files),
    }


def collect_from_refresh(tenant_id, refresh_id, store_name, columns=None, order_qty_header="Order Qty"):
    """Build the sorted, split files for a whole Refresh -> ([(name, bytes)], total)."""
    conn = get_connection()
    try:
        assignments = export_repository.all_assignment_items(conn, tenant_id, refresh_id)
    finally:
        conn.close()

    items = [{"assignment_id": a["assignment_id"], "qty": a["assigned_qty"]} for a in assignments]
    if not items:
        raise HTTPException(status_code=400, detail="No products to sort for this order.")

    files, total = docs.build_sorted_split(
        tenant_id, items, columns or [], order_qty_header, store_name, _MAX_PER_FILE
    )
    logger.info("Shelf sort tenant=%s refresh=%s products=%s files=%s",
                tenant_id, refresh_id, total, len(files))
    return files, total


# --------------------------------------------------------------------------
# Upload path — sort an existing Excel file from disk
# --------------------------------------------------------------------------

def _norm(v):
    return re.sub(r"[^a-z0-9]", "", str(v).lower()) if v is not None else ""


def _code_of(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    return s or None


def resolve_categories(tenant_id, rows):
    """rows: [{name, unit}] -> {name: category}. Resolution priority:
      1. the learned table (a human/LLM/cross-store decision for this product),
      2. the local UnitDescription + keyword rules,
      3. another store's UnitDescription (cross-store, matched by name),
      4. Others.
    """
    names = [r["name"] for r in rows if r["name"]]
    learned = pcat.get_many(tenant_id, names)  # {norm_key: {category, ...}}

    result = {}
    need_cross = []
    for r in rows:
        name = r["name"]
        key = pcat.normalize_name(name)
        if key and key in learned:
            result[name] = learned[key]["category"]
            continue
        cat = docs._detect_category(r.get("unit"), name)
        if cat != "Others":
            result[name] = cat
            continue
        need_cross.append(name)

    if need_cross:
        cross = pcat.cross_store_units(tenant_id, need_cross)  # {norm_key: unit}
        for name in need_cross:
            unit = cross.get(pcat.normalize_name(name))
            result[name] = docs._detect_category(unit, name) if unit else "Others"
    return result


def debug_peek(data):
    """A safe textual summary of an uploaded workbook (sheet names, dimensions,
    header row, first data row) for diagnostics — never raises."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
        ws = wb.active
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))] if ws else []
        first = None
        for r in ws.iter_rows(min_row=2, max_row=2, values_only=True):
            first = list(r)
            break
        return (f"sheets={wb.sheetnames} active={getattr(ws, 'title', None)} "
                f"dims={getattr(ws, 'dimensions', None)} max_row={getattr(ws, 'max_row', None)} "
                f"max_col={getattr(ws, 'max_column', None)}\nheader={header}\nfirst_row={first}")
    except Exception as e:
        return f"debug_peek failed: {e!r}"


def collect_from_file(tenant_id, store_id, store_name, data, max_per_file=_MAX_PER_FILE):
    """Sort a user-supplied Purchase Order .xlsx by shelf category and split it
    into pick-sized files, preserving the uploaded sheet's exact columns,
    values and cell formatting — only the row order and split boundaries
    change. UnitDescription/SubLocation are joined from sync.Products by the
    file's Product Code column (the sort keys are never written as columns).
    Returns ([(name, bytes)], total).
    """
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter

    try:
        wb = load_workbook(BytesIO(data), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read the uploaded file as an Excel workbook (.xlsx).")
    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=400, detail="The uploaded workbook has no active sheet.")
    max_col, max_row = ws.max_column or 0, ws.max_row or 0
    if not max_col or max_row < 2:
        raise HTTPException(status_code=400, detail="The uploaded Excel has no product rows.")

    # Find the header row and its columns. Billing-software exports often put a
    # title/date/blank rows above the header and use their own column names, so
    # we scan the first rows for a header and accept common aliases rather than
    # assuming row 1 with exact "Product Code"/"Product Name" labels.
    def _resolve(header_cells):
        labels = {}
        for c in header_cells:
            key = _norm(c.value)
            if key and key not in labels:
                labels[key] = c.column

        def pick(aliases):
            for a in aliases:
                if a in labels:
                    return labels[a]
            return None

        return {
            "code": pick(_ALIAS_CODE),
            "name": pick(_ALIAS_NAME),
            "subloc": pick(_ALIAS_SUBLOC),
            "unit": pick(_ALIAS_UNIT),
            "sno": pick(_ALIAS_SNO),
        }

    header_row, cols = None, None
    for hr in range(1, min(max_row, 20) + 1):
        candidate = _resolve(list(ws[hr]))
        if candidate["code"] or candidate["name"]:
            header_row, cols = hr, candidate
            break
    if header_row is None:
        raise HTTPException(
            status_code=400,
            detail="Couldn't find a Product/Item Code or Name column in the Excel to sort by shelf.",
        )
    code_col, name_col = cols["code"], cols["name"]
    subloc_col, unit_col, sno_col = cols["subloc"], cols["unit"], cols["sno"]

    # Only real product rows — skip blanks. An .xlsx edited/re-saved elsewhere
    # often reports a wildly inflated max_row (stray formatting to row
    # 1,048,576); without this the split would balloon into thousands of empty
    # files and exhaust memory.
    def _has_content(ri):
        for col in (name_col, code_col):
            if col:
                v = ws.cell(row=ri, column=col).value
                if v is not None and str(v).strip():
                    return True
        return False

    data_rows = [ri for ri in range(header_row + 1, max_row + 1) if _has_content(ri)]
    if not data_rows:
        raise HTTPException(status_code=400, detail="No product rows found in the uploaded Excel.")

    # Join the master by Product Code for UnitDescription + SubLocation.
    codes = set()
    if code_col:
        for ri in data_rows:
            c = _code_of(ws.cell(row=ri, column=code_col).value)
            if c:
                codes.add(c)
    master = {}
    if codes:
        conn = get_connection()
        try:
            master = export_repository.product_master(conn, tenant_id, store_id, list(codes))
        finally:
            conn.close()

    # Pass 1: name / local UnitDescription / SubLocation per row.
    rowinfo = {}
    for ri in data_rows:
        code = _code_of(ws.cell(row=ri, column=code_col).value) if code_col else None
        name = ws.cell(row=ri, column=name_col).value if name_col else None
        name = "" if name is None else str(name)
        m = master.get(code, {}) if code else {}
        unit = (ws.cell(row=ri, column=unit_col).value if unit_col else None) or m.get("unit_description")
        sub = (ws.cell(row=ri, column=subloc_col).value if subloc_col else None) or m.get("sub_location")
        rowinfo[ri] = {"name": name, "unit": unit, "sub": "" if sub is None else str(sub).strip()}

    # Pass 2: resolve categories with the learned agent (learned table ->
    # local rule/UnitDescription -> other stores' UnitDescription -> Others).
    cat_by_name = resolve_categories(tenant_id, [{"name": i["name"], "unit": i["unit"]} for i in rowinfo.values()])
    metas = {ri: (cat_by_name.get(info["name"], "Others"), info["sub"], info["name"]) for ri, info in rowinfo.items()}

    def sort_key(ri):
        cat, sub, name = metas[ri]
        return (docs._SHELF_RANK.get(cat, len(docs._SHELF_ORDER)), 1 if not sub else 0, sub, name)

    ordered = sorted(data_rows, key=sort_key)

    safe = docs._safe_name(store_name)
    files = []
    for start in range(0, len(ordered), max_per_file):
        chunk = ordered[start:start + max_per_file]
        out = Workbook()
        ows = out.active
        ows.title = ws.title or "Sheet1"

        # Column widths + hidden flags (keeps a hidden Assignment ID hidden).
        for ci in range(1, max_col + 1):
            letter = get_column_letter(ci)
            if letter in ws.column_dimensions:
                src = ws.column_dimensions[letter]
                dst = ows.column_dimensions[letter]
                if src.width:
                    dst.width = src.width
                dst.hidden = src.hidden

        def copy_row(src_idx, dst_idx):
            for ci in range(1, max_col + 1):
                s = ws.cell(row=src_idx, column=ci)
                d = ows.cell(row=dst_idx, column=ci)
                d.value = s.value
                # Copy the real style OBJECTS, never s._style: the latter is an
                # array of indices into the SOURCE workbook's style tables, and
                # reusing those indices in a new workbook makes openpyxl raise
                # IndexError on save. Assigning the objects re-registers them.
                if s.has_style:
                    try:
                        d.font = copy(s.font)
                        d.fill = copy(s.fill)
                        d.border = copy(s.border)
                        d.alignment = copy(s.alignment)
                        d.number_format = s.number_format
                        d.protection = copy(s.protection)
                    except Exception:
                        pass

        copy_row(header_row, 1)  # detected header -> row 1
        # Rightmost "Category" column: the shelf-group short code each row is
        # ordered by (TAB/OIN/RESP/…, NULL when unrecognised). Added after all
        # existing columns so nothing existing moves.
        cat_col = max_col + 1
        hdr_cat = ows.cell(row=1, column=cat_col)
        hdr_cat.value = "Category"
        hc = ws.cell(row=header_row, column=1)
        if hc.has_style:
            try:
                hdr_cat.font = copy(hc.font)
                hdr_cat.fill = copy(hc.fill)
                hdr_cat.alignment = copy(hc.alignment)
                hdr_cat.border = copy(hc.border)
            except Exception:
                pass
        ows.column_dimensions[get_column_letter(cat_col)].width = 10

        for n, ri in enumerate(chunk, start=1):
            copy_row(ri, n + 1)
            if sno_col:
                ows.cell(row=n + 1, column=sno_col).value = n
            ows.cell(row=n + 1, column=cat_col).value = docs.category_code(metas[ri][0])

        try:
            ows.protection.sheet = bool(ws.protection.sheet)
        except Exception:
            pass
        buf = BytesIO()
        out.save(buf)
        seq = start // max_per_file + 1
        files.append((f"{safe}_Sorted_{seq:02d}.xlsx", buf.getvalue()))

    logger.info("Shelf sort (upload) tenant=%s store=%s products=%s files=%s",
                tenant_id, store_id, len(ordered), len(files))
    return files, len(ordered)
