"""Export Document generation (Sprint: Export Monitor overhaul) — the
configurable per-supplier purchase order document (Excel default, or PDF /
Image), replacing the old client-side CSV-as-Excel + browser-print-as-PDF.

Excel is the working "send to supplier, get it back" sheet: Status and
Available Qty are real, sheet-protected, data-validated editable cells (every
other cell is locked) plus a hidden assignment_id column so
supplier_reply_service can match rows back to assignments on re-import even if
the supplier reorders or edits visible text. PDF/Image are print-style
snapshots of the same columns — Status/Available Qty render as blank boxes.

Row data (PTR/Cost/MRP/Offer/Discount%, product name, sub-location, unit
description) comes from assignment_repository.list_for_export — the same
PurchaseTrans-backed enrichment the Export Monitor screen itself uses.
"""

import re
from io import BytesIO

from modules.procurement import assignment_repository as repo

# Column key -> header label. Order here is the fixed left-to-right layout;
# callers select a subset of the OPTIONAL keys — mandatory ones are always on.
_MANDATORY = ["sno", "product_name", "order_qty", "mrp"]
_OPTIONAL_ORDER = ["ptr", "cost", "offer", "discount_pct", "product_code"]
_LABELS = {
    "sno": "S.No",
    "product_name": "Product Name",
    "mrp": "MRP",
    "ptr": "PTR",
    "cost": "Cost",
    "offer": "Offer",
    "discount_pct": "Dis %",
    "product_code": "Product Code",
}

# --- Shelf category detection (Shelf Sorting & Excel Split) -----------------
# Category is derived purely for row ORDERING — it is never stored, never
# added as a column, and never leaves this module. Primary signal is the
# product's UnitDescription; when that is blank it is inferred from the
# ProductName. See _detect_category.

# The picking sort order the store follows down the shelves. Medicine forms
# first, then oral-care / topical, then nutrition / supplies, then anything
# unrecognised.
_SHELF_ORDER = [
    "Tablet", "Capsule", "Syrup", "Respule", "Rotacap", "Inhaler",
    "Injection", "Drops", "Cream", "Ointment", "Gel", "Lotion",
    "Spray", "Paste", "Powder", "Sachet", "Wash", "Soap", "Food",
    "Bandage", "Cloth", "Veterinary", "Others",
]
_SHELF_RANK = {c: i for i, c in enumerate(_SHELF_ORDER)}

# Keyword -> category. Order matters: name-strong / non-drug groups first (so a
# toothpaste or nutrition product is not mis-filed by a drug-form word), then
# the drug forms. Plurals are handled (RESPULES, ROTACAPS, TABS…). Word
# boundaries keep e.g. \bCAP\b from matching inside "ROTACAP".
_CATEGORY_PATTERNS = [
    ("Veterinary", re.compile(r"\b(?:VET|VETY|VETERINARY)\b", re.I)),
    ("Food", re.compile(r"\b(?:ENSURE|PROTINEX|PEDIASURE|LACTARE|PROTEIN|NUTRITION|"
                        r"GRANULES?|MALT|HORLICKS|BOURNVITA|COMPLAN|PRO\s?PL|PROPL)\b", re.I)),
    ("Soap", re.compile(r"\bSOAP\b", re.I)),
    ("Wash", re.compile(r"\b(?:BODYWASH|FACEWASH|HANDWASH|WASH|SHAMPOO)\b", re.I)),
    ("Paste", re.compile(r"\b(?:TOOTHPASTE|PASTE)\b", re.I)),
    ("Spray", re.compile(r"\bSPRAYS?\b", re.I)),
    ("Sachet", re.compile(r"\bSACHETS?\b", re.I)),
    ("Powder", re.compile(r"\bPOWDERS?\b", re.I)),
    ("Bandage", re.compile(r"\b(?:BANDAGE|GAUZE|CREPE|BAND\s?AID|DRESSING)\b", re.I)),
    ("Cloth", re.compile(r"\bCLOTH\b", re.I)),
    ("Rotacap", re.compile(r"\bROTA(?:CAPS?)?\b", re.I)),
    ("Respule", re.compile(r"\bRESP(?:ULES?)?\b", re.I)),
    ("Tablet", re.compile(r"\bTAB(?:LETS?|S)?\b", re.I)),
    ("Capsule", re.compile(r"\b(?:CAPS?(?:ULES?)?|TRANSCAPS?)\b", re.I)),
    ("Syrup", re.compile(r"\b(?:SYRUPS?|SYP|SYR)\b", re.I)),
    ("Inhaler", re.compile(r"\b(?:INHALERS?|INH)\b", re.I)),
    ("Injection", re.compile(r"\b(?:INJECTIONS?|INJ)\b", re.I)),
    ("Drops", re.compile(r"\bDROPS?\b", re.I)),
    ("Cream", re.compile(r"\bCREAMS?\b", re.I)),
    ("Ointment", re.compile(r"\b(?:OINTMENTS?|OINTM|OINT|OIN)\b", re.I)),
    ("Gel", re.compile(r"\bGELS?\b", re.I)),
    ("Lotion", re.compile(r"\bLOTIONS?\b", re.I)),
]

# Name-strong groups identified from the NAME *first*: these items may still
# carry a drug-form UnitDescription (TAB/CAP) that would otherwise mis-file
# them, so a name match wins over UnitDescription.
_SPECIAL = {"Veterinary", "Food", "Soap", "Wash", "Paste", "Spray",
            "Sachet", "Powder", "Bandage", "Cloth"}
_SPECIAL_PATTERNS = [(c, p) for c, p in _CATEGORY_PATTERNS if c in _SPECIAL]

# Short code shown in the optional "Category" column (rightmost). Unrecognised
# rows show NULL, matching the store's own wording.
_CATEGORY_CODE = {
    "Tablet": "TAB", "Capsule": "CAP", "Syrup": "SYP", "Respule": "RESP",
    "Rotacap": "ROTA", "Inhaler": "INH", "Injection": "INJ", "Drops": "DROP",
    "Cream": "CREAM", "Ointment": "OIN", "Gel": "GEL", "Lotion": "LOTION",
    "Spray": "SPRAY", "Paste": "PASTE", "Powder": "PWD", "Sachet": "SACHET",
    "Wash": "WASH", "Soap": "SOAP", "Food": "FOOD",
    "Bandage": "BAND", "Cloth": "CLOTH", "Veterinary": "VET",
    "Others": "NULL",
}


def category_code(category):
    return _CATEGORY_CODE.get(category or "Others", "NULL")


def _match_category(text, patterns=_CATEGORY_PATTERNS):
    if not text:
        return None
    for cat, pat in patterns:
        if pat.search(text):
            return cat
    return None


def _detect_category(unit_description, product_name):
    """Category used only for shelf ordering. Non-drug/general groups
    (veterinary, soap, bandage, cloth) are recognised from the ProductName
    first — they must not be scattered into Tablet/Capsule just because their
    UnitDescription says TAB/CAP. Otherwise UnitDescription is the primary
    signal, falling back to the ProductName; anything unrecognised -> 'Others'."""
    name = product_name or ""
    special = _match_category(name, _SPECIAL_PATTERNS)
    if special:
        return special
    ud = (unit_description or "").strip()
    cat = _match_category(ud) if ud else None
    if cat is None:
        cat = _match_category(name)
    return cat or "Others"


def _safe_name(store_name):
    base = re.sub(r"[^A-Za-z0-9._-]+", "_", (store_name or "Store").strip()).strip("_")
    return base or "Store"


def _sort_key(row, sort_by):
    if sort_by == "shelf":
        # Category order, then SubLocation (NULL/blank goes last), then name.
        cat = row.get("_category") or "Others"
        sub = (row.get("sub_location") or "").strip()
        return (
            _SHELF_RANK.get(cat, len(_SHELF_ORDER)),
            1 if not sub else 0,
            sub,
            row.get("product_name") or "",
        )
    if sort_by == "sub_location":
        return (row.get("sub_location") or "", row.get("product_name") or "")
    if sort_by == "unit_description":
        return (row.get("unit_description") or "", row.get("product_name") or "")
    return (row.get("product_name") or "", row.get("product_code") or "")


def _cell(row, key, sno, qty_by_id):
    if key == "sno":
        return sno
    if key == "product_name":
        return row.get("product_name") or row.get("product_code") or "—"
    if key == "order_qty":
        return qty_by_id.get(row["assignment_id"], row.get("assigned_qty") or 0)
    if key == "mrp":
        return row.get("pt_mrp")
    if key == "ptr":
        return row.get("pt_ptr")
    if key == "cost":
        return row.get("pt_cost")
    if key == "offer":
        return row.get("pt_offer") or ""
    if key == "discount_pct":
        return row.get("pt_discount_pct")
    if key == "product_code":
        return row.get("product_code")
    return None


def _build_rows(tenant_id, items, columns, order_qty_header, sort_by):
    """items: [{assignment_id, qty}] -> (column plan, sorted row dicts) shared
    by every output format. Column plan is [(key, label), ...] in display
    order: the 4 mandatory columns first, then any selected optional ones."""
    ids = [i["assignment_id"] for i in items]
    qty_by_id = {i["assignment_id"]: i["qty"] for i in items}
    data = repo.list_for_export(tenant_id, ids)
    for r in data:
        r["_category"] = _detect_category(r.get("unit_description"), r.get("product_name"))
    data.sort(key=lambda r: _sort_key(r, sort_by))

    plan = []
    for key in _MANDATORY:
        label = (order_qty_header or "").strip() or "Order Qty" if key == "order_qty" else _LABELS[key]
        plan.append((key, label))
    for key in _OPTIONAL_ORDER:
        if key in columns:
            plan.append((key, _LABELS[key]))

    out_rows = []
    for i, row in enumerate(data, start=1):
        out_rows.append({
            "assignment_id": row["assignment_id"],
            "values": [_cell(row, key, i, qty_by_id) for key, _ in plan],
            "category": row.get("_category"),  # for the optional Category column
        })
    return plan, out_rows


def build_excel(tenant_id, items, columns, order_qty_header, sort_by):
    plan, rows = _build_rows(tenant_id, items, columns, order_qty_header, sort_by)
    return _excel_from_plan_rows(plan, rows)


def _excel_from_plan_rows(plan, rows):
    """The single source of the Purchase Order .xlsx layout — the exact same
    workbook (columns, sheet protection, green Order Qty fill, hidden
    assignment_id, data validation, autosize) whether it is a normal export or
    one split of the Shelf Sorting output. Callers pass a pre-built (plan,
    rows) so the format never diverges between the two paths."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Protection
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    headers = [label for _, label in plan] + ["Status", "Available Qty"]
    order_qty_col = next(i for i, (key, _) in enumerate(plan) if key == "order_qty") + 1
    status_col = len(plan) + 1
    avail_col = len(plan) + 2
    hidden_col = len(plan) + 3  # assignment_id — round-trip key, not for display

    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Order"

    ws.append(headers + ["Assignment ID"])
    for c in ws[1]:
        c.font = Font(bold=True)

    green = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
    for r in rows:
        ws.append(r["values"] + ["", "", r["assignment_id"]])

    last_row = len(rows) + 1

    # Sheet protection: everything locked except the Status/Available Qty
    # cells the supplier is meant to fill in — enforced by Excel itself, not
    # just a convention. Header row stays fully locked.
    ws.protection.sheet = True
    unlocked = Protection(locked=False)
    locked = Protection(locked=True)
    for row_idx in range(1, last_row + 1):
        for col_idx in range(1, hidden_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            editable = row_idx > 1 and col_idx in (status_col, avail_col)
            cell.protection = unlocked if editable else locked
            if row_idx > 1 and col_idx == order_qty_col:
                cell.fill = green

    dv = DataValidation(type="list", formula1='"Available,Partial,Not Available"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(status_col)}2:{get_column_letter(status_col)}{last_row}")

    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row_idx in range(2, last_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            max_len = max(max_len, len(str(v)) if v is not None else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

    ws.column_dimensions[get_column_letter(hidden_col)].hidden = True

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf(tenant_id, items, columns, order_qty_header, sort_by):
    import fitz  # pymupdf

    plan, rows = _build_rows(tenant_id, items, columns, order_qty_header, sort_by)
    headers = [label for _, label in plan] + ["Status", "Available Qty"]

    doc = fitz.open()
    page_w, page_h = fitz.paper_size("a4")
    margin = 24
    row_h = 16
    usable_w = page_w - 2 * margin
    col_w = usable_w / len(headers)
    rows_per_page = int((page_h - 2 * margin) // row_h) - 1

    def new_page():
        page = doc.new_page(width=page_w, height=page_h)
        y = margin
        for i, h in enumerate(headers):
            x = margin + i * col_w
            page.draw_rect(fitz.Rect(x, y, x + col_w, y + row_h), color=(0, 0, 0), width=0.5)
            page.insert_text((x + 3, y + row_h - 4), str(h)[:22], fontsize=7.5, fontname="helv")
        return page, y + row_h

    page, y = new_page()
    on_page = 0
    for r in rows:
        if on_page >= rows_per_page:
            page, y = new_page()
            on_page = 0
        values = [str(v) if v is not None else "" for v in r["values"]] + ["", ""]
        for i, v in enumerate(values):
            x = margin + i * col_w
            page.draw_rect(fitz.Rect(x, y, x + col_w, y + row_h), color=(0, 0, 0), width=0.3)
            page.insert_text((x + 3, y + row_h - 4), v[:22], fontsize=7, fontname="helv")
        y += row_h
        on_page += 1

    buf = BytesIO()
    doc.save(buf)
    doc.close()
    return buf.getvalue()


def build_image(tenant_id, items, columns, order_qty_header, sort_by):
    from PIL import Image, ImageDraw, ImageFont

    plan, rows = _build_rows(tenant_id, items, columns, order_qty_header, sort_by)
    headers = [label for _, label in plan] + ["Status", "Available Qty"]
    order_qty_idx = next(i for i, (key, _) in enumerate(plan) if key == "order_qty")

    col_w, row_h = 140, 26
    width = col_w * len(headers)
    height = row_h * (len(rows) + 1)
    img = Image.new("RGB", (max(width, 1), max(height, row_h)), "white")
    draw = ImageDraw.Draw(img)
    try:
        font = ImageFont.truetype("arial.ttf", 13)
        font_bold = ImageFont.truetype("arialbd.ttf", 13)
    except OSError:
        font = ImageFont.load_default()
        font_bold = font

    def draw_row(y, values, bold, green_col=None):
        for i, v in enumerate(values):
            x = i * col_w
            if green_col is not None and i == green_col:
                draw.rectangle([x, y, x + col_w, y + row_h], fill="#C6EFCE")
            draw.rectangle([x, y, x + col_w, y + row_h], outline="black")
            draw.text((x + 4, y + 6), str(v)[:20], fill="black", font=font_bold if bold else font)

    draw_row(0, headers, bold=True)
    for r_idx, r in enumerate(rows, start=1):
        values = [v if v is not None else "" for v in r["values"]] + ["", ""]
        draw_row(r_idx * row_h, values, bold=False, green_col=order_qty_idx)

    buf = BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


_BUILDERS = {"excel": build_excel, "pdf": build_pdf, "image": build_image}
_MEDIA_TYPES = {
    "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
    "image": "image/png",
}
_EXTENSIONS = {"excel": "xlsx", "pdf": "pdf", "image": "png"}


def build_document(tenant_id, items, fmt, columns, order_qty_header, sort_by, supplier_code=None):
    """Returns (bytes, filename, media_type) for the requested format."""
    builder = _BUILDERS.get(fmt, build_excel)
    content = builder(tenant_id, items, columns or [], order_qty_header, sort_by)
    from datetime import date
    name_bits = ["PO", supplier_code or "export", date.today().isoformat()]
    filename = "_".join(str(b) for b in name_bits if b) + f".{_EXTENSIONS[fmt]}"
    return content, filename, _MEDIA_TYPES[fmt]


def build_sorted_split(tenant_id, items, columns, order_qty_header, store_name, max_per_file=16):
    """Shelf Sorting & Excel Split: take the existing export dataset, sort it
    by shelf category -> SubLocation -> ProductName, then slice it into
    consecutive .xlsx files of at most `max_per_file` products each. Every file
    is the identical Purchase Order layout (same columns, header, formatting);
    only row order and the split boundaries differ. S.No restarts at 1 in each
    file. A rightmost "Category" column shows the shelf group's short code
    (TAB/OIN/RESP/…, NULL when unrecognised). Returns ([(filename, bytes),
    ...], total_products)."""
    plan, rows = _build_rows(tenant_id, items, columns, order_qty_header, "shelf")
    plan = plan + [("category", "Category")]  # rightmost order-value column
    sno_idx = next((i for i, (k, _) in enumerate(plan) if k == "sno"), None)
    safe = _safe_name(store_name)

    files = []
    for start in range(0, len(rows), max_per_file):
        chunk = rows[start:start + max_per_file]
        renumbered = []
        for n, r in enumerate(chunk, start=1):
            values = list(r["values"]) + [category_code(r.get("category"))]
            if sno_idx is not None:
                values[sno_idx] = n
            renumbered.append({"assignment_id": r["assignment_id"], "values": values})
        seq = start // max_per_file + 1
        content = _excel_from_plan_rows(plan, renumbered)
        files.append((f"{safe}_Sorted_{seq:02d}.xlsx", content))
    return files, len(rows)
