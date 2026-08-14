"""Internal Supplier Stock Distribution — pipeline + config + logs.

Generates one Excel file per target store from a single source ("HO") store's
own stock — e.g. NMW acting as an internal supplier for NMS/NMC/NMG/NMA — and
writes the results into the SAME procurement.supplier_stock table the
external Supplier Live Stock importer uses (supplier_stock_import.py),
tagged supplier_code = the target store's own local code for the source
(procurement.store_supplier_map, mirrors legacy Ho_code), source =
'internal_distribution'. That keeps every downstream read
(GET .../supplier-stock) unchanged: HO shows up exactly like any other
supplier.

Rack/SubLocation: the source provider (distribution_providers.py) reads NMW's
own PRODUCTS.SubLocation and carries it through as `rack` on every row. This
mirrors the already-proven legacy port (modules/legacy_order/repository.py
replace_supplier_stock), including its carry-forward rule: a delete+reinsert
must not blank out a rack that was already on file just because this cycle's
source row happens to have no SubLocation set.

Pipeline per run: collect source stock ONCE -> validate source store belongs
to the tenant -> for each enabled target store (own transaction boundary):
validate supplier mapping -> replace that store's procurement.supplier_stock
rows for this internal supplier (STOCK stage) -> generate + save Excel
(EXCEL stage) -> send via WhatsApp (WHATSAPP stage, never rolls back STOCK or
EXCEL on failure) -> log all three stage statuses separately. One store's
failure never stops the others.
"""

import os
import time
import uuid

from config.database import get_connection
from modules.procurement import supplier_stock_import as ssi
from modules.procurement.distribution_providers import get_provider
from modules.procurement.reports.excel_exporter import ExcelExporter
from modules.procurement._dbutil import as_uid, rows_to_dicts, stringify

_SQL_DIR = os.path.join(os.path.dirname(__file__), "sql")
_DDL_FILES = [
    os.path.join(_SQL_DIR, "0021_supplier_stock_distribution.sql"),
    os.path.join(_SQL_DIR, "0022_store_supplier_map.sql"),
    os.path.join(_SQL_DIR, "0023_supplier_stock_distribution_rack_status.sql"),
]
_DEFAULT_STORAGE_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "storage", "distribution"
)


def _storage_dir():
    """Export directory — configurable via DISTRIBUTION_EXPORT_PATH (server
    config, not a client-supplied path) so ops can point it at a shared/UNC
    location without touching frontend code. Falls back to the existing
    backend/storage/distribution folder."""
    path = os.getenv("DISTRIBUTION_EXPORT_PATH", "").strip() or _DEFAULT_STORAGE_DIR
    os.makedirs(path, exist_ok=True)
    if not os.access(path, os.W_OK):
        raise RuntimeError(f"Distribution export path is not writable: {path}")
    return path


def apply_ddl():
    conn = get_connection()
    try:
        cur = conn.cursor()
        for ddl_file in _DDL_FILES:
            with open(ddl_file, "r", encoding="utf-8") as fh:
                script = fh.read()
            for batch in [b.strip() for b in script.split("\nGO") if b.strip()]:
                cur.execute(batch)
        conn.commit()
    finally:
        conn.close()


def _stores(conn, tenant_id):
    cur = conn.cursor()
    cur.execute(
        "SELECT store_id, store_code, store_name FROM dbo.stores WHERE tenant_id = ?",
        (tenant_id,),
    )
    return rows_to_dicts(cur)


# --------------------------------------------------------------------------
# Config (which stores receive the feed + where to notify)
# --------------------------------------------------------------------------

def list_config(tenant_id, source_store_code="NMW"):
    apply_ddl()
    conn = get_connection()
    try:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT s.store_id, s.store_code, s.store_name,
                   s.w_group_name AS whatsapp_group, c.phone_number,
                   COALESCE(c.enabled, 1) AS enabled,
                   m.local_supplier_code
            FROM dbo.stores s
            LEFT JOIN procurement.distribution_config c
                ON c.tenant_id = s.tenant_id AND c.store_id = s.store_id
            LEFT JOIN procurement.store_supplier_map m
                ON m.tenant_id = s.tenant_id AND m.store_id = s.store_id
               AND m.source_store_code = ?
            WHERE s.tenant_id = ?
            ORDER BY s.store_code
            """,
            (source_store_code, tenant_id),
        )
        return [stringify(r) for r in rows_to_dicts(cur)]
    finally:
        conn.close()


# --------------------------------------------------------------------------
# Store supplier code mapping -- what a target store's OWN order/purchase
# screens call the source store as a supplier (mirrors legacy Ho_code:
# NMS/NMA='94', NMC='ST_2', NMG='99'). Falls back to the source store code
# itself when a store has no mapping row.
# --------------------------------------------------------------------------

def local_supplier_code(conn, tenant_id, store_id, source_store_code):
    cur = conn.cursor()
    cur.execute(
        "SELECT local_supplier_code FROM procurement.store_supplier_map "
        "WHERE tenant_id = ? AND store_id = ? AND source_store_code = ?",
        (tenant_id, store_id, source_store_code),
    )
    row = cur.fetchone()
    return row[0] if row else source_store_code


def require_local_supplier_code(conn, tenant_id, store_id, source_store_code):
    """Same lookup as ``local_supplier_code`` but with NO silent fallback --
    used by the STOCK-update stage so an unmapped store fails loudly (per
    distribution requirements) instead of writing rows under a supplier code
    that store's own order screens don't actually recognise as this source."""
    cur = conn.cursor()
    cur.execute(
        "SELECT local_supplier_code FROM procurement.store_supplier_map "
        "WHERE tenant_id = ? AND store_id = ? AND source_store_code = ?",
        (tenant_id, store_id, source_store_code),
    )
    row = cur.fetchone()
    return row[0] if row else None


def import_legacy_supplier_map(tenant_id, source_store_code="NMW"):
    """One-shot: copy dbo.Stores.Ho_code from the legacy OrderNMC database
    into procurement.store_supplier_map, keyed by store_code. Stores with no
    Ho_code set (or that don't exist in this tenant) are skipped."""
    from modules.legacy_order import repository as legacy_repo

    legacy_stores = {
        s["store_name"]: s["ho_code"]
        for s in legacy_repo.list_stores(active_only=False)
        if s["ho_code"] and s["store_name"] != source_store_code
    }
    conn = get_connection()
    try:
        stores = {s["store_code"]: s["store_id"] for s in _stores(conn, tenant_id)}
    finally:
        conn.close()

    imported, skipped = [], []
    for store_code, ho_code in legacy_stores.items():
        store_id = stores.get(store_code)
        if not store_id:
            skipped.append(store_code)
            continue
        save_supplier_map(tenant_id, store_id, source_store_code, ho_code)
        imported.append({"store_code": store_code, "local_supplier_code": ho_code})
    return {"imported": imported, "skipped": skipped}


def save_supplier_map(tenant_id, store_id, source_store_code, local_code):
    apply_ddl()
    conn = get_connection()
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT store_supplier_map_id FROM procurement.store_supplier_map "
            "WHERE tenant_id = ? AND store_id = ? AND source_store_code = ?",
            (tenant_id, store_id, source_store_code),
        )
        row = cur.fetchone()
        if row:
            cur.execute(
                "UPDATE procurement.store_supplier_map SET local_supplier_code = ?, "
                "updated_at = GETDATE() WHERE store_supplier_map_id = ?",
                (local_code, row[0]),
            )
        else:
            cur.execute(
                "INSERT INTO procurement.store_supplier_map "
                "(tenant_id, store_id, source_store_code, local_supplier_code) "
                "VALUES (?, ?, ?, ?)",
                (tenant_id, store_id, source_store_code, local_code),
            )
        conn.commit()
        return {"saved": True}
    finally:
        conn.close()


def save_config(tenant_id, store_id, whatsapp_group, phone_number, enabled):
    """WhatsApp group name lives on dbo.stores.w_group_name (a property of
    the store, not of one distribution feed); phone_number/enabled stay in
    procurement.distribution_config."""
    apply_ddl()
    conn = get_connection()
    try:
        cur = conn.cursor()
        cur.execute(
            "UPDATE dbo.stores SET w_group_name = ?, updated_at = GETDATE() "
            "WHERE tenant_id = ? AND store_id = ?",
            (whatsapp_group, tenant_id, store_id),
        )

        cur.execute(
            "SELECT distribution_config_id FROM procurement.distribution_config "
            "WHERE tenant_id = ? AND store_id = ?",
            (tenant_id, store_id),
        )
        row = cur.fetchone()
        if row:
            cur.execute(
                "UPDATE procurement.distribution_config SET phone_number = ?, "
                "enabled = ?, updated_at = GETDATE() WHERE distribution_config_id = ?",
                (phone_number, 1 if enabled else 0, row[0]),
            )
        else:
            cur.execute(
                "INSERT INTO procurement.distribution_config "
                "(tenant_id, store_id, store_code, phone_number, enabled) "
                "SELECT ?, ?, store_code, ?, ? FROM dbo.stores WHERE store_id = ?",
                (tenant_id, store_id, phone_number, 1 if enabled else 0, store_id),
            )
        conn.commit()
        return {"saved": True}
    finally:
        conn.close()


# --------------------------------------------------------------------------
# Generation pipeline
# --------------------------------------------------------------------------

def _target_stores(conn, tenant_id, source_store_code, only_store_ids=None):
    stores = _stores(conn, tenant_id)
    targets = [s for s in stores if s["store_code"] != source_store_code]
    if only_store_ids:
        wanted = {str(i) for i in only_store_ids}
        targets = [s for s in targets if str(s["store_id"]) in wanted]
    cur = conn.cursor()
    cur.execute(
        "SELECT store_id, enabled FROM procurement.distribution_config WHERE tenant_id = ?",
        (tenant_id,),
    )
    enabled_map = {str(r[0]): bool(r[1]) for r in cur.fetchall()}
    return [t for t in targets if enabled_map.get(str(t["store_id"]), True)]


def _build_workbook(rows, target_store_code):
    sheet_rows = [
        {
            "ProductCode": r["code"],
            "ProductName": r["name"],
            "stock": r["stock"],
            "ProductDiscPercent": r["disc_percent"],
            "SubLocation": r.get("rack") or "",
        }
        for r in rows
    ]
    return ExcelExporter.build_workbook(sheet_rows, f"SupplierStock_{target_store_code}")


def _replace_supplier_stock(conn, tenant_id, target_store_id, supplier_code, rows):
    """Legacy-compatible replace: delete + reinsert this supplier's rows for
    the target store, but carry a product's previously-recorded rack forward
    when this cycle's source row has none — mirrors
    modules/legacy_order/repository.py replace_supplier_stock exactly, just
    against procurement.supplier_stock instead of the central legacy table."""
    cur = conn.cursor()
    cur.execute(
        "SELECT supplier_product_code, rack FROM procurement.supplier_stock "
        "WHERE tenant_id = ? AND store_id = ? AND supplier_code = ? "
        "AND source = 'internal_distribution' AND rack IS NOT NULL",
        (tenant_id, target_store_id, supplier_code),
    )
    existing_racks = {str(r[0]): r[1] for r in cur.fetchall()}

    cur.execute(
        "DELETE FROM procurement.supplier_stock "
        "WHERE tenant_id = ? AND store_id = ? AND supplier_code = ? AND source = 'internal_distribution'",
        (tenant_id, target_store_id, supplier_code),
    )
    stock_rows = [(
        tenant_id, target_store_id, supplier_code, r["code"], r["name"],
        None, r["stock"], None, None, r["disc_percent"], None, None, None,
        None, None, "internal_distribution", 1, None,
        r.get("rack") or existing_racks.get(str(r["code"])),
    ) for r in rows]
    inserted = ssi.insert_stock_rows(conn, stock_rows)
    ssi.resolve_product_codes(conn, tenant_id, target_store_id, supplier_code)
    return inserted


def _send_whatsapp(conn, run_item_id, tenant_id, target_store_id, excel_path, message):
    """Send the just-generated Excel file to the target store's configured
    WhatsApp group/number and record the outcome. Runs strictly AFTER the
    stock-update commit, in its own try/except in the caller, so a WhatsApp
    failure here can never roll back or be confused with a stock-update
    failure (see distribution_run_item.stock_status vs whatsapp_status)."""
    cur = conn.cursor()
    cur.execute(
        "SELECT s.w_group_name, c.phone_number, COALESCE(c.enabled, 1) "
        "FROM dbo.stores s "
        "LEFT JOIN procurement.distribution_config c ON c.tenant_id = s.tenant_id AND c.store_id = s.store_id "
        "WHERE s.tenant_id = ? AND s.store_id = ?",
        (tenant_id, target_store_id),
    )
    row = cur.fetchone()
    if not row or not row[2] or not (row[0] or row[1]):
        return "not_queued", None

    whatsapp_group, phone_number = row[0], row[1]
    cur.execute(
        "INSERT INTO procurement.distribution_queue "
        "(run_item_id, tenant_id, store_id, whatsapp_group, phone_number, excel_path, status) "
        "VALUES (?, ?, ?, ?, ?, ?, 'queued')",
        (run_item_id, tenant_id, target_store_id, whatsapp_group, phone_number, excel_path),
    )
    conn.commit()

    try:
        from modules.whatsapp import service as wa
        profiles = wa._load_profiles()
        profile = next((p for p in profiles if p["is_default"]), profiles[0] if profiles else None)
        if not profile:
            raise RuntimeError("No WhatsApp profile configured to send from.")

        with open(excel_path, "rb") as fh:
            attachment_content = fh.read()
        attachment_name = os.path.basename(excel_path)

        if whatsapp_group:
            targets = wa._load_targets()
            target = next(
                (t for t in targets if t["profile_id"] == profile["profile_id"]
                 and t["target_kind"] == "group" and t["target_name"] == whatsapp_group),
                None,
            )
            if target is None:
                created = wa.upsert_target({
                    "profile_id": profile["profile_id"],
                    "target_kind": "group",
                    "target_name": whatsapp_group,
                    "target_ref": whatsapp_group,
                    "can_send": True,
                })
                target = created["target"]
            result = wa.send_target_message(target["target_id"], message, attachment_name, attachment_content)
        else:
            result = wa.send_message(profile["profile_id"], phone_number, message, attachment_name, attachment_content)

        status = "sent" if result.get("status") in ("sent",) else (
            "queued" if result.get("status") in ("manual", "manual-fallback", "launched") else "failed"
        )
        error = None if status != "failed" else result.get("message")
        cur.execute(
            "UPDATE procurement.distribution_queue SET status = ?, sent_at = CASE WHEN ? = 'sent' THEN GETDATE() ELSE sent_at END, "
            "error_message = ? WHERE run_item_id = ?",
            (status, status, error, run_item_id),
        )
        conn.commit()
        return status, error
    except Exception as exc:
        cur.execute(
            "UPDATE procurement.distribution_queue SET status = 'failed', error_message = ? WHERE run_item_id = ?",
            (str(exc)[:1000], run_item_id),
        )
        conn.commit()
        return "failed", str(exc)[:1000]


def _sync_source_store(source_store_code):
    """Block until the source store's branch data is freshly synced into the
    central OrderNMC mirror (modules.legacy_order.service.start_sync) --
    required before every export per owner instruction: read current stock
    only after a completed sync, never a stale/mid-sync snapshot. Only
    meaningful for the legacy branch-DB source; raises on failure so the
    caller aborts the run rather than exporting against an unknown state."""
    from modules.legacy_order import service as legacy_service

    job_id = legacy_service.start_sync(source_store_code)
    while True:
        job = legacy_service.get_job(job_id)
        if not job or job["status"] != "running":
            break
        time.sleep(2)
    if not job or job["status"] != "completed":
        message = (job or {}).get("message", "unknown error")
        raise RuntimeError(f"NMW sync did not complete: {message}")
    return job


def _source_store_in_tenant(conn, tenant_id, source_store_code):
    cur = conn.cursor()
    cur.execute(
        "SELECT 1 FROM dbo.stores WHERE tenant_id = ? AND store_code = ?",
        (tenant_id, source_store_code),
    )
    return cur.fetchone() is not None


def generate(tenant_id, source_store_code, provider_name, only_store_ids=None,
             started_by=None, excel_only=False, supplier_update_only=False):
    """Run the full distribution pipeline.

    ``excel_only``: STOCK stage skipped entirely (no supplier_stock write);
    ``supplier_update_only``: EXCEL + WHATSAPP stages skipped entirely (no
    file written, nothing sent). Otherwise all three stages run for every
    enabled target store, each with its own status so a WhatsApp failure
    never reads as a stock-update failure and vice versa.
    """
    apply_ddl()
    storage_dir = _storage_dir()
    provider = get_provider(provider_name)
    run_stamp = time.strftime("%Y%m%d_%H%M%S")
    run_date = time.strftime("%Y%m%d")

    conn = get_connection()
    try:
        if not _source_store_in_tenant(conn, tenant_id, source_store_code):
            return {"run_id": None, "error": f"Source store '{source_store_code}' does not "
                    "belong to the active tenant.", "items": []}

        run_id = str(uuid.uuid4())
        targets = _target_stores(conn, tenant_id, source_store_code, only_store_ids)
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO procurement.distribution_run "
            "(run_id, tenant_id, source_store_code, provider, stores_total, started_by) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (run_id, tenant_id, source_store_code, provider_name, len(targets), as_uid(started_by)),
        )
        conn.commit()

        if provider_name == "legacy":
            try:
                _sync_source_store(source_store_code)
            except Exception as exc:
                cur.execute(
                    "UPDATE procurement.distribution_run SET status = 'failed', "
                    "finished_at = GETDATE(), error_summary = ? WHERE run_id = ?",
                    (f"NMW sync failed: {exc}"[:2000], run_id),
                )
                conn.commit()
                return {"run_id": run_id, "error": f"NMW sync failed: {exc}", "items": []}

        try:
            source_rows = provider.fetch_stock(source_store_code)
        except Exception as exc:
            cur.execute(
                "UPDATE procurement.distribution_run SET status = 'failed', "
                "finished_at = GETDATE(), error_summary = ? WHERE run_id = ?",
                (f"Could not read source stock: {exc}"[:2000], run_id),
            )
            conn.commit()
            return {"run_id": run_id, "error": f"Could not read source stock: {exc}", "items": []}

        total_stock_qty = sum(r["stock"] for r in source_rows)
        cur.execute(
            "UPDATE procurement.distribution_run SET total_products = ?, total_stock_qty = ? WHERE run_id = ?",
            (len(source_rows), total_stock_qty, run_id),
        )
        conn.commit()

        succeeded, failed, items, errors = 0, 0, [], []
        for store in targets:
            t0 = time.time()
            run_item_id = str(uuid.uuid4())
            cur.execute(
                "INSERT INTO procurement.distribution_run_item "
                "(run_item_id, run_id, tenant_id, store_id, store_code) VALUES (?, ?, ?, ?, ?)",
                (run_item_id, run_id, tenant_id, store["store_id"], store["store_code"]),
            )
            conn.commit()

            stock_status, stock_error = "skipped", None
            excel_status, excel_error = "skipped", None
            wa_status, wa_error = "not_queued", None
            excel_path, imported, item_failed = None, 0, False

            # --- STOCK stage -------------------------------------------------
            if not excel_only:
                try:
                    supplier_code = require_local_supplier_code(conn, tenant_id, store["store_id"], source_store_code)
                    if not supplier_code:
                        raise ValueError(
                            f"No local supplier code mapped for {store['store_code']} "
                            f"(source {source_store_code}) — configure it before running."
                        )
                    imported = _replace_supplier_stock(
                        conn, tenant_id, store["store_id"], supplier_code, source_rows
                    )
                    conn.commit()
                    stock_status = "success"
                except Exception as exc:
                    conn.rollback()
                    stock_status, stock_error, item_failed = "failed", str(exc)[:1000], True

            # --- EXCEL stage ---------------------------------------------------
            if not supplier_update_only:
                try:
                    content = _build_workbook(source_rows, store["store_code"])
                    filename = f"{source_store_code}_Stock_{store['store_code']}_{run_date}.xlsx"
                    excel_path = os.path.join(storage_dir, filename)
                    with open(excel_path, "wb") as fh:
                        fh.write(content)
                    excel_status = "success"
                except Exception as exc:
                    excel_status, excel_error, item_failed = "failed", str(exc)[:1000], True
                    excel_path = None

            # --- WHATSAPP stage (never rolls back STOCK/EXCEL on failure) -----
            # Only runs on the full pipeline ("Generate All"/"Generate
            # Selected") -- "Generate Excel Only" must not send, "Update
            # Supplier Stock Only" must not export or send.
            if excel_path and excel_status == "success" and not supplier_update_only and not excel_only:
                message = f"{source_store_code} Supplier Stock — {store['store_code']} ({run_stamp})"
                wa_status, wa_error = _send_whatsapp(
                    conn, run_item_id, tenant_id, store["store_id"], excel_path, message
                )

            duration_ms = int((time.time() - t0) * 1000)
            overall_status = "failed" if item_failed else "success"
            cur.execute(
                "UPDATE procurement.distribution_run_item SET finished_at = GETDATE(), "
                "duration_ms = ?, rows_exported = ?, rows_imported = ?, excel_path = ?, "
                "supplier_updated = ?, whatsapp_status = ?, whatsapp_error = ?, "
                "stock_status = ?, stock_error = ?, excel_status = ?, excel_error = ?, "
                "status = ?, error_message = ? WHERE run_item_id = ?",
                (duration_ms, len(source_rows), imported, excel_path,
                 1 if stock_status == "success" else 0, wa_status, wa_error,
                 stock_status, stock_error, excel_status, excel_error,
                 overall_status, stock_error or excel_error, run_item_id),
            )
            conn.commit()

            if item_failed:
                failed += 1
                errors.append(f"{store['store_code']}: {stock_error or excel_error}")
            else:
                succeeded += 1
            items.append({
                "store_code": store["store_code"], "status": overall_status,
                "rows": len(source_rows), "stock_status": stock_status,
                "excel_status": excel_status, "whatsapp_status": wa_status,
                "error": stock_error or excel_error,
            })

        cur.execute(
            "UPDATE procurement.distribution_run SET status = 'completed', finished_at = GETDATE(), "
            "stores_succeeded = ?, stores_failed = ?, error_summary = ? WHERE run_id = ?",
            (succeeded, failed, ("; ".join(errors))[:2000] if errors else None, run_id),
        )
        conn.commit()
        return {"run_id": run_id, "stores_total": len(targets),
                "stores_succeeded": succeeded, "stores_failed": failed, "items": items}
    finally:
        conn.close()


def list_runs(tenant_id, limit=20):
    apply_ddl()
    conn = get_connection()
    try:
        cur = conn.cursor()
        cur.execute(
            f"""
            SELECT TOP ({int(limit)}) run_id, source_store_code, provider, started_at, finished_at,
                   status, stores_total, stores_succeeded, stores_failed,
                   total_products, total_stock_qty, error_summary
            FROM procurement.distribution_run
            WHERE tenant_id = ?
            ORDER BY started_at DESC
            """,
            (tenant_id,),
        )
        return [stringify(r) for r in rows_to_dicts(cur)]
    finally:
        conn.close()


def run_detail(run_id):
    conn = get_connection()
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT * FROM procurement.distribution_run WHERE run_id = ?",
            (run_id,),
        )
        run = rows_to_dicts(cur)
        cur.execute(
            "SELECT * FROM procurement.distribution_run_item WHERE run_id = ? ORDER BY store_code",
            (run_id,),
        )
        items = rows_to_dicts(cur)
        return {
            "run": stringify(run[0]) if run else None,
            "items": [stringify(i) for i in items],
        }
    finally:
        conn.close()


def run_item_products(run_item_id):
    """Read back the product rows from the already-generated Excel file for
    one run item (for the WhatsApp image / preview) -- a pure read of a file
    the generation pipeline already wrote; does not touch stock, rack or
    distribution logic."""
    from openpyxl import load_workbook

    conn = get_connection()
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT store_code, excel_path FROM procurement.distribution_run_item WHERE run_item_id = ?",
            (run_item_id,),
        )
        row = cur.fetchone()
        if not row or not row[1] or not os.path.exists(row[1]):
            return {"store_code": row[0] if row else None, "excel_path": None, "rows": []}
        store_code, excel_path = row
        wb = load_workbook(excel_path, read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, [])
        header_idx = {h: i for i, h in enumerate(headers)}
        rows = []
        for r in rows_iter:
            if not r or r[0] is None:
                continue
            rows.append({
                "code": r[header_idx.get("ProductCode", 0)],
                "name": r[header_idx.get("ProductName", 1)],
                "stock": r[header_idx.get("stock", 2)],
                "rack": r[header_idx.get("SubLocation")] if "SubLocation" in header_idx else None,
            })
        return {"store_code": store_code, "excel_path": excel_path, "rows": rows}
    finally:
        conn.close()


def retry_failed(run_id, provider_name=None, started_by=None):
    conn = get_connection()
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT tenant_id, source_store_code, provider FROM procurement.distribution_run WHERE run_id = ?",
            (run_id,),
        )
        run = cur.fetchone()
        if not run:
            raise ValueError("Run not found")
        tenant_id, source_store_code, provider = run
        cur.execute(
            "SELECT store_id FROM procurement.distribution_run_item WHERE run_id = ? AND status = 'failed'",
            (run_id,),
        )
        failed_store_ids = [str(r[0]) for r in cur.fetchall()]
        if not failed_store_ids:
            return {"run_id": run_id, "items": [], "stores_total": 0, "stores_succeeded": 0, "stores_failed": 0}
        return generate(
            str(tenant_id), source_store_code, provider_name or provider,
            only_store_ids=failed_store_ids, started_by=started_by,
        )
    finally:
        conn.close()
