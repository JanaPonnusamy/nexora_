from io import BytesIO

import xlwt
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    @classmethod
    def auto_width(cls, ws):
        for column in ws.columns:
            length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    length = max(length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[column_name].width = length + 5

    @classmethod
    def format_header(cls, ws):
        for cell in ws[1]:
            cell.fill = cls.HEADER_FILL
            cell.font = cls.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = cls.BORDER

    @classmethod
    def apply_filters(cls, ws):
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

    @classmethod
    def apply_number_formats(cls, ws):
        currency_columns = [
            "Sales",
            "Purchase",
            "PurchaseReturn",
            "OpeningStock",
            "TransferIn",
            "TransferOut",
            "Adjustment",
            "ClosingStock",
            "CostOfSales",
            "PendingAmount",
            "GP",
        ]
        percent_columns = ["GPPercent", "PurchaseSalesRatio", "StockPendingRatio"]
        header_map = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}

        for header in currency_columns:
            if header not in header_map:
                continue
            for row in range(2, ws.max_row + 1):
                ws.cell(row, header_map[header]).number_format = "#,##0.00"

        for header in percent_columns:
            if header not in header_map:
                continue
            for row in range(2, ws.max_row + 1):
                ws.cell(row, header_map[header]).number_format = "0.00"

    @classmethod
    def add_totals(cls, ws):
        if ws.max_row < 2:
            return
        total_row = ws.max_row + 1
        ws.cell(total_row, 1).value = "TOTAL"
        for col in range(2, ws.max_column + 1):
            column_letter = get_column_letter(col)
            ws.cell(total_row, col).value = (
                f"=SUM({column_letter}2:{column_letter}{total_row - 1})"
            )

    @classmethod
    def conditional_formatting(cls, ws):
        for col in range(1, ws.max_column + 1):
            if ws.cell(1, col).value == "GP":
                letter = get_column_letter(col)
                ws.conditional_formatting.add(
                    f"{letter}2:{letter}{ws.max_row}",
                    CellIsRule(operator="lessThan", formula=["0"]),
                )

    @classmethod
    def build_workbook(cls, rows, title):
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]

        if rows:
            headers = list(rows[0].keys())
            ws.append(headers)
            for row in rows:
                ws.append([row.get(header) for header in headers])
        else:
            ws.append(["No Data"])

        cls.format_header(ws)
        cls.apply_filters(ws)
        cls.apply_number_formats(ws)
        cls.add_totals(ws)
        cls.auto_width(ws)
        cls.conditional_formatting(ws)

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @classmethod
    def build_workbook_xls(cls, rows, title):
        """Legacy Excel 97-2003 (.xls, BIFF) output — some 3rd-party apps
        that receive these exports can't open openpyxl's .xlsx (OOXML)."""
        wb = xlwt.Workbook(encoding="utf-8")
        ws = wb.add_sheet(title[:31] or "Sheet1")

        header_style = xlwt.easyxf(
            "font: bold on, color white; "
            "pattern: pattern solid, fore_colour dark_blue; "
            "align: horiz center; "
            "borders: left thin, right thin, top thin, bottom thin;"
        )
        cell_style = xlwt.easyxf("borders: left thin, right thin, top thin, bottom thin;")

        if rows:
            headers = list(rows[0].keys())
            widths = [len(str(h)) for h in headers]
            for col, header in enumerate(headers):
                ws.write(0, col, header, header_style)
            for row_idx, row in enumerate(rows, start=1):
                for col, header in enumerate(headers):
                    value = row.get(header)
                    ws.write(row_idx, col, value, cell_style)
                    widths[col] = max(widths[col], len(str(value)) if value is not None else 0)
            for col, width in enumerate(widths):
                ws.col(col).width = min((width + 5) * 256, 65535)
        else:
            ws.write(0, 0, "No Data")

        ws.set_panes_frozen(True)
        ws.set_horz_split_pos(1)

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()
