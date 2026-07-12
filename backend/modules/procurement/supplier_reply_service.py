"""Supplier Reply import service (Sprint: Export Monitor overhaul).

The supplier fills in Status (Available / Partial / Not Available) and
Available Qty on the Excel export_document_service generated, and sends it
back. This reads it, matches rows to assignments via the hidden Assignment ID
column, and records a PRE-SHIPMENT confirmation — never a real GRN receipt
(reconciliation_service owns that state machine). A partial/not_available
reply rolls the owning order item's remaining_qty up so it surfaces in the
existing Pending tab with zero Pending-module changes.
"""

from io import BytesIO

from fastapi import HTTPException

from config.database import get_connection
from modules.procurement import supplier_reply_repository as repo

_VALID_STATUS = {"available", "partial", "not_available"}


def _normalize_status(raw):
    s = str(raw).strip().lower().replace(" ", "_") if raw not in (None, "") else ""
    return s if s in _VALID_STATUS else None


def preview_reply_import(tenant_id, file_bytes, filename):
    """Parse the uploaded reply Excel: headers, matched rows, warnings for
    anything that no longer lines up with a live assignment."""
    from openpyxl import load_workbook

    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception:
        raise HTTPException(
            status_code=400,
            detail="Could not read the uploaded file as an Excel workbook",
        )
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    try:
        status_idx = headers.index("Status")
        avail_idx = headers.index("Available Qty")
        id_idx = headers.index("Assignment ID")
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail="This doesn't look like an exported Purchase Order sheet "
            "(missing Status / Available Qty / Assignment ID columns)",
        )

    parsed = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        assignment_id = r[id_idx] if id_idx < len(r) else None
        if not assignment_id:
            continue
        status_raw = r[status_idx] if status_idx < len(r) else None
        avail_qty = r[avail_idx] if avail_idx < len(r) else None
        parsed.append({
            "assignment_id": str(assignment_id).strip(),
            "status_raw": status_raw,
            "status": _normalize_status(status_raw),
            "available_qty": float(avail_qty) if avail_qty not in (None, "") else None,
        })

    live = {
        a["assignment_id"]: a
        for a in repo.get_assignments_by_ids(tenant_id, [p["assignment_id"] for p in parsed])
    }

    rows = []
    for p in parsed:
        a = live.get(p["assignment_id"])
        warning = None
        if not a:
            warning = "No matching assignment found — row skipped (renamed/removed since export)."
        elif p["status"] is None and p["status_raw"] not in (None, ""):
            warning = f'Unrecognized status "{p["status_raw"]}" — row skipped.'
        rows.append({
            "assignment_id": p["assignment_id"],
            "product_name": a.get("product_name") if a else None,
            "product_code": a.get("product_code") if a else None,
            "supplier_code": a.get("supplier_code") if a else None,
            "assigned_qty": a.get("assigned_qty") if a else None,
            "status": p["status"],
            "available_qty": p["available_qty"],
            "warning": warning,
            "applicable": bool(a and p["status"]),
        })
    return {"filename": filename, "rows": rows}


def apply_reply_import(tenant_id, rows, imported_by):
    conn = get_connection()
    applied = 0
    skipped = 0
    try:
        for row in rows:
            status = _normalize_status(row.get("status"))
            assignment_id = row.get("assignment_id")
            if not assignment_id or not status:
                skipped += 1
                continue
            assignment = repo.get_assignment(conn, tenant_id, assignment_id)
            if not assignment:
                skipped += 1
                continue
            assigned_qty = float(assignment.get("assigned_qty") or 0)
            if status == "available":
                reply_qty = assigned_qty
            elif status == "not_available":
                reply_qty = 0.0
            else:
                reply_qty = min(assigned_qty, max(0.0, float(row.get("available_qty") or 0)))
            repo.apply_reply(conn, tenant_id, assignment_id, status, reply_qty, imported_by)
            if status in ("partial", "not_available"):
                shortfall = max(0, assigned_qty - reply_qty)
                repo.mark_shortfall(conn, tenant_id, assignment["order_item_id"], shortfall)
            applied += 1
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return {"applied": applied, "skipped": skipped}
