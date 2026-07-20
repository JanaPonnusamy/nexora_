"""Excel preview/import helpers for procurement.supplier_stock.

This importer is platform-only: it writes to NEXORA_PLATFORM through
config.database and does not read any legacy database.
"""

from io import BytesIO

from openpyxl import load_workbook

from modules.supplier_stock_analysis import repository

TARGETS = [
    {"value": "supplier_product_code", "label": "Supplier Product Code", "mandatory": True},
    {"value": "supplier_product_name", "label": "Supplier Product Name", "mandatory": True},
    {"value": "available_stock", "label": "Stock / Qty", "mandatory": True},
    {"value": "ptr", "label": "PTR", "mandatory": False},
    {"value": "mrp", "label": "MRP", "mandatory": False},
    {"value": "discount", "label": "Discount", "mandatory": False},
    {"value": "packing", "label": "Packing", "mandatory": False},
    {"value": "scheme", "label": "Scheme", "mandatory": False},
    {"value": "free", "label": "Free Qty", "mandatory": False},
    {"value": "minimum_qty", "label": "Min Qty", "mandatory": False},
    {"value": "transaction_date", "label": "Transaction Date", "mandatory": False},
]

_SUGGESTIONS = {
    "code": "supplier_product_code",
    "productcode": "supplier_product_code",
    "itemcode": "supplier_product_code",
    "name": "supplier_product_name",
    "productname": "supplier_product_name",
    "itemname": "supplier_product_name",
    "stock": "available_stock",
    "qty": "available_stock",
    "quantity": "available_stock",
    "ptr": "ptr",
    "rate": "ptr",
    "mrp": "mrp",
    "discount": "discount",
    "disc": "discount",
    "packing": "packing",
    "pack": "packing",
    "scheme": "scheme",
    "sch": "scheme",
    "free": "free",
    "minqty": "minimum_qty",
    "minimumqty": "minimum_qty",
}


def _norm(value):
    return "".join(ch for ch in str(value or "").lower() if ch.isalnum())


def _number(value):
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _sheet(file_bytes):
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    return workbook.active


def preview(file_bytes):
    sheet = _sheet(file_bytes)
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
    headers = [str(h or "").strip() for h in header_row]
    suggested = {}
    for header in headers:
        key = _norm(header)
        suggested[header] = _SUGGESTIONS.get(key, "")
    sample = []
    for row in sheet.iter_rows(min_row=2, max_row=min(sheet.max_row, 11), values_only=True):
        sample.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return {
        "sheet_name": sheet.title,
        "row_count": max(sheet.max_row - 1, 0),
        "headers": headers,
        "suggested_mapping": suggested,
        "targets": TARGETS,
        "sample_rows": sample,
    }


def import_file(tenant_id, store_id, supplier_code, file_bytes, mapping, imported_by=None):
    reverse = {target: source for source, target in mapping.items() if target}
    missing = [
        t["value"] for t in TARGETS
        if t["mandatory"] and t["value"] not in reverse
    ]
    if missing:
        return {"success": False, "error": "Missing required mapping", "missing": missing}

    sheet = _sheet(file_bytes)
    headers = [str(h or "").strip() for h in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])]
    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        raw = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        product_code = raw.get(reverse["supplier_product_code"])
        product_name = raw.get(reverse["supplier_product_name"])
        if product_code in (None, "") or product_name in (None, ""):
            continue
        rows.append({
            "supplier_product_code": str(product_code).strip(),
            "supplier_product_name": str(product_name).strip(),
            "available_stock": _number(raw.get(reverse["available_stock"])),
            "ptr": _number(raw.get(reverse.get("ptr"))),
            "mrp": _number(raw.get(reverse.get("mrp"))),
            "discount": _number(raw.get(reverse.get("discount"))),
            "packing": raw.get(reverse.get("packing")),
            "scheme": raw.get(reverse.get("scheme")),
            "free": _number(raw.get(reverse.get("free"))),
            "minimum_qty": _number(raw.get(reverse.get("minimum_qty"))),
            "transaction_date": raw.get(reverse.get("transaction_date")),
        })

    result = repository.replace_supplier_stock(
        tenant_id, store_id, supplier_code, rows, imported_by or "desktop-import"
    )
    return {"success": True, **result}
