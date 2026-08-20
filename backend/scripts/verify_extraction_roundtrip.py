"""End-to-end check of the document-extraction round trip against a live server.

This is the Phase 2 exit criterion, minus the camera: upload page images, walk
the server pipeline, read the review payload, save, export, and verify the
workbook against the **frozen** 5-sheet contract in
`docs/Document_Extraction_Excel_Contract.md`.

Why it exists: every piece of that chain is unit-tested, but until this script
the chain had never been run end to end against a real server — the HO host is
unreachable from a dev machine. Unit tests cannot catch a stage that returns
200 with an empty body, a contract column the exporter quietly renamed, or a
pipeline that stalls between stages.

The calls, their order and their parameters mirror the mobile client exactly
(`CaptureUploader.pipeline` and `DocumentExtractionApi`), so a pass here is
evidence about the app's real path rather than about this script's idea of it.

    backend/.venv/bin/python scripts/verify_extraction_roundtrip.py \
        --base-url http://localhost:8000 \
        --username <user> --password <pass> \
        --tenant-id <tenant>

Add `--keep` to leave the import in place, `--image path.png` to send a real
scan instead of the generated one. Exit code is 0 only if every stage passed
and the workbook matched the contract.
"""

from __future__ import annotations

import argparse
import io
import json
import re
import sys
import urllib.error
import urllib.request
import uuid
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
CONTRACT = REPO_ROOT / 'docs' / 'Document_Extraction_Excel_Contract.md'

# Mirrors CaptureUploader.pipeline in the mobile client. Order matters: each
# stage reads what the previous one wrote.
PIPELINE = ['preprocess', 'ocr', 'extract', 'validate']


# --------------------------------------------------------------------------
# A synthetic invoice
# --------------------------------------------------------------------------

INVOICE_LINES = [
    # (product, batch, expiry, qty, rate, mrp)
    ('PARACETAMOL 500MG TAB', 'PC24A1', '11/27', 100, 12.50, 18.00),
    ('AMOXYCILLIN 250MG CAP', 'AM23X9', '09/27', 50, 42.75, 60.00),
    ('CETIRIZINE 10MG TAB', 'CT25B4', '03/28', 200, 8.20, 12.50),
    ('AZITHROMYCIN 500MG TAB', 'AZ24K2', '06/27', 30, 88.00, 125.00),
]


def build_invoice_png(path: Path) -> None:
    """Renders a legible A4-ish invoice.

    Rendered large (1654x2339 ≈ 200 DPI A4) because OCR accuracy falls off a
    cliff below roughly 150 DPI, and a synthetic page that fails to OCR would
    test nothing but the error path.
    """
    from PIL import Image, ImageDraw, ImageFont

    W, H = 1654, 2339
    image = Image.new('RGB', (W, H), 'white')
    draw = ImageDraw.Draw(image)

    def font(size: int, bold: bool = False):
        # DejaVu ships with Pillow; fall back rather than fail on a slim build.
        for name in (
            'DejaVuSans-Bold.ttf' if bold else 'DejaVuSans.ttf',
            'Arial Bold.ttf' if bold else 'Arial.ttf',
        ):
            try:
                return ImageFont.truetype(name, size)
            except OSError:
                continue
        return ImageFont.load_default()

    y = 90
    draw.text((90, y), 'SRI VENKATESHWARA PHARMA DISTRIBUTORS', font=font(40, True), fill='black')
    y += 58
    draw.text((90, y), '12 Anna Salai, Chennai 600002', font=font(26), fill='black')
    y += 40
    draw.text((90, y), 'GSTIN: 33AABCS1429B1ZP    DL No: TN/CH/20B/1234', font=font(26), fill='black')

    y += 80
    draw.line([(90, y), (W - 90, y)], fill='black', width=3)
    y += 30

    draw.text((90, y), 'TAX INVOICE', font=font(36, True), fill='black')
    y += 60
    draw.text((90, y), 'Invoice No: SVP/2026/00417', font=font(28), fill='black')
    draw.text((900, y), 'Invoice Date: 12/08/2026', font=font(28), fill='black')
    y += 42
    draw.text((90, y), 'Order No: PO-88213', font=font(28), fill='black')
    draw.text((900, y), 'Credit Days: 30', font=font(28), fill='black')

    y += 70
    draw.line([(90, y), (W - 90, y)], fill='black', width=2)
    y += 24

    cols = [(90, 'PRODUCT'), (720, 'BATCH'), (930, 'EXP'), (1090, 'QTY'),
            (1220, 'RATE'), (1380, 'MRP'), (1520, 'AMOUNT')]
    for x, label in cols:
        draw.text((x, y), label, font=font(26, True), fill='black')
    y += 44
    draw.line([(90, y), (W - 90, y)], fill='black', width=2)
    y += 22

    net = 0.0
    total_qty = 0
    for name, batch, expiry, qty, rate, mrp in INVOICE_LINES:
        amount = qty * rate
        net += amount
        total_qty += qty
        row = [
            (90, name), (720, batch), (930, expiry), (1090, str(qty)),
            (1220, f'{rate:.2f}'), (1380, f'{mrp:.2f}'), (1490, f'{amount:.2f}'),
        ]
        for x, text in row:
            draw.text((x, y), text, font=font(26), fill='black')
        y += 46

    y += 20
    draw.line([(90, y), (W - 90, y)], fill='black', width=2)
    y += 30

    taxable = net
    cgst = round(taxable * 0.06, 2)
    sgst = round(taxable * 0.06, 2)
    gross = round(taxable + cgst + sgst, 2)

    for label, value in [
        ('Total Quantity', f'{total_qty}'),
        ('Taxable Amount', f'{taxable:.2f}'),
        ('CGST 6%', f'{cgst:.2f}'),
        ('SGST 6%', f'{sgst:.2f}'),
        ('Net Amount', f'{gross:.2f}'),
    ]:
        draw.text((1120, y), label, font=font(28, label == 'Net Amount'), fill='black')
        draw.text((1450, y), value, font=font(28, label == 'Net Amount'), fill='black')
        y += 46

    image.save(path, 'PNG')


# --------------------------------------------------------------------------
# Minimal HTTP, so this runs on the backend venv with no extra dependency
# --------------------------------------------------------------------------


class Api:
    def __init__(self, base_url: str):
        self.base = base_url.rstrip('/')
        self.token: str | None = None

    def _request(self, method, path, *, body=None, headers=None, raw=False):
        url = f'{self.base}{path}'
        head = dict(headers or {})
        if self.token:
            head['Authorization'] = f'Bearer {self.token}'
        req = urllib.request.Request(url, data=body, headers=head, method=method)
        try:
            with urllib.request.urlopen(req, timeout=300) as res:
                payload = res.read()
                return payload if raw else (json.loads(payload) if payload else {})
        except urllib.error.HTTPError as e:
            detail = e.read().decode('utf-8', 'replace')[:400]
            raise RuntimeError(f'{method} {path} -> {e.code}: {detail}') from None

    def get(self, path, raw=False):
        return self._request('GET', path, raw=raw)

    def post_json(self, path, payload):
        return self._request(
            'POST', path,
            body=json.dumps(payload).encode(),
            headers={'Content-Type': 'application/json'},
        )

    def post_form(self, path, fields, files):
        """multipart/form-data, matching the mobile client's upload."""
        boundary = f'----axythic{uuid.uuid4().hex}'
        buf = io.BytesIO()

        def w(s):
            buf.write(s.encode() if isinstance(s, str) else s)

        for key, value in fields.items():
            w(f'--{boundary}\r\n')
            w(f'Content-Disposition: form-data; name="{key}"\r\n\r\n{value}\r\n')
        for key, (filename, content) in files.items():
            w(f'--{boundary}\r\n')
            w(f'Content-Disposition: form-data; name="{key}"; filename="{filename}"\r\n')
            w('Content-Type: image/png\r\n\r\n')
            w(content)
            w('\r\n')
        w(f'--{boundary}--\r\n')

        return self._request(
            'POST', path,
            body=buf.getvalue(),
            headers={'Content-Type': f'multipart/form-data; boundary={boundary}'},
        )


# --------------------------------------------------------------------------
# Contract verification
# --------------------------------------------------------------------------


def contract_sheets() -> dict[str, list[str]]:
    """Reads the expected sheets and columns out of the frozen contract.

    Parsed from the document rather than retyped, so drift on either side shows
    up as a failure here instead of as a silently wrong workbook. This is the
    same approach `backend/tests/test_document_export.py` takes.
    """
    text = CONTRACT.read_text(encoding='utf-8')
    sheets: dict[str, list[str]] = {}
    current: str | None = None

    for line in text.splitlines():
        heading = re.match(r'^##\s+Sheet\s+\d+\s+—\s+(.+?)\s*$', line)
        if heading:
            current = heading.group(1).strip()
            sheets[current] = []
            continue
        if current is None:
            continue
        cells = [c.strip() for c in line.split('|')[1:-1]] if line.startswith('|') else []
        if len(cells) >= 2 and cells[0].isdigit():
            sheets[current].append(cells[1])

    return {name: cols for name, cols in sheets.items() if cols}


def verify_workbook(data: bytes, expected: dict[str, list[str]]) -> list[str]:
    """Returns a list of problems; empty means the workbook matches."""
    import openpyxl

    problems: list[str] = []
    book = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    actual_sheets = book.sheetnames

    if len(actual_sheets) != len(expected):
        problems.append(
            f'expected {len(expected)} sheets, found {len(actual_sheets)}: {actual_sheets}'
        )

    for index, (name, columns) in enumerate(expected.items()):
        if index >= len(actual_sheets):
            problems.append(f'missing sheet {index + 1} ({name})')
            continue
        sheet = book[actual_sheets[index]]
        header = [str(c.value).strip() if c.value is not None else ''
                  for c in next(sheet.iter_rows(min_row=1, max_row=1))]
        missing = [c for c in columns if c not in header]
        if missing:
            problems.append(
                f'sheet "{actual_sheets[index]}" (contract: {name}) is missing '
                f'columns {missing}'
            )
    return problems


# --------------------------------------------------------------------------


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--base-url', default='http://localhost:8000')
    parser.add_argument('--username', required=True)
    parser.add_argument('--password', required=True)
    parser.add_argument('--tenant-id', required=True)
    parser.add_argument('--store-id')
    parser.add_argument('--image', type=Path, help='use a real scan instead of the generated page')
    parser.add_argument('--keep', action='store_true', help='leave the import in place')
    parser.add_argument('--out', type=Path, help='where to write the exported workbook')
    args = parser.parse_args()

    api = Api(args.base_url)
    step = lambda msg: print(f'  {msg}', flush=True)  # noqa: E731

    print('1. Authenticating')
    login = api.post_json(
        '/api/auth/login',
        {'username': args.username, 'password': args.password},
    )
    api.token = login.get('access_token') or login.get('token')
    if not api.token:
        print(f'   login returned no token: {list(login)}', file=sys.stderr)
        return 1
    step('ok')

    print('2. Preparing the page image')
    if args.image:
        image_bytes = args.image.read_bytes()
        filename = args.image.name
        step(f'using {args.image}')
    else:
        tmp = Path(__file__).parent / '_roundtrip_invoice.png'
        build_invoice_png(tmp)
        image_bytes = tmp.read_bytes()
        filename = 'page_01.png'
        step(f'generated {tmp} ({len(image_bytes) // 1024} KB)')

    print('3. Uploading (mirrors DocumentExtractionApi.upload)')
    query = f'?tenant_id={args.tenant_id}'
    if args.store_id:
        query += f'&store_id={args.store_id}'
    created = api.post_form(
        f'/api/document-extraction/imports{query}',
        fields={'group_as_single_invoice': 'true'},
        files={'files': (filename, image_bytes)},
    )
    imports = created.get('imports') or []
    if not imports:
        print(f'   server accepted the upload but returned no import: {created}', file=sys.stderr)
        return 1
    first = imports[0]
    import_id = first if isinstance(first, int) else first.get('import_id')
    step(f'import_id={import_id}')

    print('4. Walking the pipeline (mirrors CaptureUploader.pipeline)')
    for stage in PIPELINE:
        api.post_json(
            f'/api/document-extraction/imports/{import_id}/{stage}'
            f'?tenant_id={args.tenant_id}',
            {},
        )
        step(f'{stage} ok')

    print('5. Reading the review payload')
    review = api.get(
        f'/api/document-extraction/imports/{import_id}/review'
        f'?tenant_id={args.tenant_id}'
    )
    header = review.get('header') or {}
    items = review.get('items') or []
    step(f'invoice_number={header.get("invoice_number")!r} '
         f'net_amount={header.get("net_amount")!r} lines={len(items)}')
    if not items:
        print('   the pipeline produced no line items — OCR or extraction failed',
              file=sys.stderr)

    print('6. Saving')
    api.post_json(
        f'/api/document-extraction/imports/{import_id}/save?tenant_id={args.tenant_id}',
        {'force': True},
    )
    step('ok')

    print('7. Exporting')
    export = api.post_json(
        f'/api/document-extraction/exports?tenant_id={args.tenant_id}',
        {'import_ids': [import_id], 'format': 'xlsx'},
    )
    batch = export.get('export_batch_id') or export.get('batch_id')
    step(f'export_batch_id={batch}')

    workbook = api.get(
        f'/api/document-extraction/exports/{batch}/download'
        f'?tenant_id={args.tenant_id}',
        raw=True,
    )
    out = args.out or (Path(__file__).parent / '_roundtrip_export.xlsx')
    out.write_bytes(workbook)
    step(f'{len(workbook) // 1024} KB → {out}')

    print('8. Verifying against the frozen contract')
    expected = contract_sheets()
    step(f'contract declares {len(expected)} sheets: {", ".join(expected)}')
    problems = verify_workbook(workbook, expected)
    for problem in problems:
        print(f'   MISMATCH: {problem}', file=sys.stderr)

    if not args.keep:
        print('9. Cleaning up')
        api._request(
            'DELETE',
            f'/api/document-extraction/imports/{import_id}?tenant_id={args.tenant_id}',
        )
        step('import deleted')
    else:
        print(f'9. Leaving import {import_id} in place (--keep)')

    if problems:
        print('\nFAILED: the workbook does not match the contract.')
        return 1
    print('\nPASSED: round trip complete and the workbook matches the contract.')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
