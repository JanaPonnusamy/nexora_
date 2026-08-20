"""Updates SubLocation on the REAL NMA branch database, from an Excel file.

Reads ProductCode + ProposedSubLocation columns (produced by
nmc_to_nma_sublocation_source.py) and applies them directly to NMA's own
store server -- not the central OrderNMC mirror, which gets overwritten on
every Sync click and is therefore not a durable place to fix this.

Touches ONLY dbo.Products.SubLocation, and ONLY for rows whose SubLocation is
currently NULL/blank/the literal text 'NULL' on the branch at execution time
(re-checked live, not trusted from the spreadsheet). No columns are added, no
other tables or columns are written, no schema changes of any kind.

Branch connection details (ServerName/Database/UserName/Password) are read
from dbo.Stores in the central OrderNMC database at runtime -- not
hardcoded -- via the existing get_branch_connection() helper.

Usage:
    backend/.venv/Scripts/python backend/scripts/nma_branch_sublocation_update_from_excel.py <source.xlsx>            # dry run (default, no writes)
    backend/.venv/Scripts/python backend/scripts/nma_branch_sublocation_update_from_excel.py <source.xlsx> --execute  # perform the guarded UPDATE inside a transaction
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook

from modules.legacy_order.database import get_central_connection, get_branch_connection

STORE_NAME = "NMA"


def _blank(s):
    return s is None or not str(s).strip() or str(s).strip().upper() == "NULL"


def read_excel_pairs(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header_row = None
    pairs = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if header_row is None:
            if row and "ProductCode" in row:
                header_row = [str(c) if c is not None else "" for c in row]
            continue
        rowmap = dict(zip(header_row, row))
        pc = rowmap.get("ProductCode")
        subloc = rowmap.get("SubLocation")
        if pc is None or _blank(subloc):
            continue
        pairs.append((int(pc), str(subloc).strip()))
    return pairs


def get_branch_credentials(store_name):
    conn = get_central_connection()
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT ServerName, [Database], UserName, Password FROM dbo.Stores WHERE StoreName = ?",
            store_name,
        )
        row = cur.fetchone()
        if not row:
            raise RuntimeError(f"No dbo.Stores row found for StoreName='{store_name}'")
        return {"server": row[0], "database": row[1], "username": row[2], "password": row[3]}
    finally:
        conn.close()


def _is_oin_unit(unit_description):
    return (unit_description or "").strip().upper() == "OIN"


def dry_run(branch_conn, pairs):
    cur = branch_conn.cursor()
    eligible = []       # (pc, name, old_subloc, old_unitdesc, proposed_subloc)
    already_set = []
    wrong_unit = []      # UnitDescription on the branch isn't 'OIN' -- never touched
    not_found = []
    for pc, subloc in pairs:
        cur.execute(
            "SELECT ProductCode, ProductName, SubLocation, UnitDescription FROM Products WHERE ProductCode = ?",
            pc,
        )
        row = cur.fetchone()
        if not row:
            not_found.append((pc, subloc))
            continue
        _, name, old_subloc, old_unitdesc = row
        if not _is_oin_unit(old_unitdesc):
            wrong_unit.append((pc, name, old_subloc, old_unitdesc, subloc))
        elif _blank(old_subloc):
            eligible.append((pc, name, old_subloc, old_unitdesc, subloc))
        else:
            already_set.append((pc, name, old_subloc, old_unitdesc, subloc))

    print("=== DRY RUN: NMA branch SubLocation update (from Excel) ===\n")
    print(f"Total pairs read from Excel:            {len(pairs)}")
    print(f"Eligible for update (UnitDescription='OIN', blank SubLocation): {len(eligible)}")
    print(f"Already set on branch (skipped):        {len(already_set)}")
    print(f"UnitDescription != 'OIN' (skipped):      {len(wrong_unit)}")
    print(f"ProductCode not found on branch:         {len(not_found)}")

    if eligible:
        print("\n--- Eligible rows (ProductCode | ProductName | OldSubLocation | OldUnitDescription -> ProposedSubLocation) ---")
        for pc, name, old_subloc, old_unitdesc, subloc in eligible:
            print(f"  {pc} | {name} | old_subloc={old_subloc!r} | unitdesc={old_unitdesc!r} -> proposed={subloc!r}")
    if already_set:
        print(f"\n--- Already set on branch, skipped (first 20 of {len(already_set)}) ---")
        for pc, name, old_subloc, old_unitdesc, subloc in already_set[:20]:
            print(f"  {pc} | {name} | current={old_subloc!r} | unitdesc={old_unitdesc!r} (proposed was {subloc!r})")
    if wrong_unit:
        print(f"\n--- UnitDescription != 'OIN', skipped (first 20 of {len(wrong_unit)}) ---")
        for pc, name, old_subloc, old_unitdesc, subloc in wrong_unit[:20]:
            print(f"  {pc} | {name} | unitdesc={old_unitdesc!r} (proposed was {subloc!r})")
    if not_found:
        print("\n--- Not found on branch (skipped) ---")
        for pc, subloc in not_found:
            print(f"  {pc} -> {subloc!r}")

    return [(pc, name, old_subloc, subloc) for pc, name, old_subloc, old_unitdesc, subloc in eligible]


def execute_update(branch_conn, eligible):
    # pyodbc connections default to autocommit=False, meaning pyodbc itself
    # already wraps everything from here to the next commit()/rollback() in
    # one transaction. Issuing raw "BEGIN/COMMIT TRANSACTION" SQL text on top
    # of that only opens/closes a NESTED SQL Server transaction -- the outer
    # pyodbc-level transaction is untouched and, if never explicitly
    # committed via branch_conn.commit(), gets silently rolled back when the
    # connection closes. Use pyodbc's own commit()/rollback(), not SQL text.
    cur = branch_conn.cursor()
    try:
        updated = []
        for pc, name, current, subloc in eligible:
            cur.execute(
                """
                UPDATE Products
                SET SubLocation = ?
                OUTPUT inserted.ProductCode, inserted.ProductName, deleted.SubLocation AS OldSubLocation, inserted.SubLocation AS NewSubLocation
                WHERE ProductCode = ?
                  AND LTRIM(RTRIM(ISNULL(SubLocation, ''))) IN ('', 'NULL')
                  AND UPPER(LTRIM(RTRIM(ISNULL(UnitDescription, '')))) = 'OIN'
                """,
                subloc, pc,
            )
            row = cur.fetchone()
            if row:
                updated.append(row)

        print(f"\n=== UPDATE executed inside transaction: {len(updated)} row(s) affected ===")

        if len(updated) != len(eligible):
            print(f"MISMATCH: expected {len(eligible)} updates, got {len(updated)}. "
                  "Rolling back -- no changes committed.")
            branch_conn.rollback()
            return None

        pcs = ",".join(str(r[0]) for r in updated) or "-1"
        cur.execute(
            f"SELECT COUNT(*) FROM Products WHERE ProductCode IN ({pcs}) "
            "AND LTRIM(RTRIM(ISNULL(SubLocation, ''))) IN ('', 'NULL')"
        )
        still_blank = cur.fetchone()[0]
        if still_blank != 0:
            print(f"VERIFICATION FAILED: {still_blank} updated rows still blank. Rolling back.")
            branch_conn.rollback()
            return None

        print("Verification passed.")
        branch_conn.commit()
        print("COMMITTED.")

        print("\n--- Updated NMA (branch) rows ---")
        for pc, name, old, new in updated:
            print(f"{pc} | {name} | {old!r} -> {new!r}")

        return updated
    except Exception:
        branch_conn.rollback()
        raise


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if a != "--execute"]
    execute = "--execute" in sys.argv
    if not args:
        print("Usage: nma_branch_sublocation_update_from_excel.py <source.xlsx> [--execute]")
        sys.exit(1)

    pairs = read_excel_pairs(args[0])
    creds = get_branch_credentials(STORE_NAME)
    print(f"Connecting to NMA branch: {creds['server']} / {creds['database']}")
    branch_conn = get_branch_connection(creds["server"], creds["database"], creds["username"], creds["password"])
    try:
        eligible = dry_run(branch_conn, pairs)
        if execute:
            if eligible:
                execute_update(branch_conn, eligible)
            else:
                print("\nNothing eligible -- skipping UPDATE.")
        else:
            print("\n(dry run only -- pass --execute to perform the guarded UPDATE)")
    finally:
        branch_conn.close()
