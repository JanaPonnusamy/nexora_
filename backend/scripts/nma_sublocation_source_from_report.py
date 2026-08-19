"""Builds the NMA SubLocation update source Excel from the existing report.

Reads scratch_nmc_oin_report.xlsx (produced by nmc_oin_sublocation_report.py)
instead of re-querying the mapping from scratch. For NMC source rows where
dbo.SupplierProductMatch found no NMA mapping, this adds a second pass:
normalize both NMC and NMA product names (uppercase, strip punctuation,
sort tokens so word order doesn't matter) and look for a unique normalized
match against NMA's live product catalog. Any new match found this way is
included and tagged NAME_MATCH so it's visibly distinct from the
SupplierProductMatch-based SUPPLIER_MATCH rows.

Only products whose NMC SubLocation is the canonical "OIN <first letter of
product name>" pattern are considered (same quality gate as
nmc_to_nma_sublocation_source.py) -- this avoids re-introducing the
EPTOIN/ISOTROIN/JOINTACE-style false positives.

Read-only against dbo.Products (NMA catalog fetch only, for the name-match
pass). Writes nothing to any database -- only the output Excel file.

Usage:
    backend/.venv/Scripts/python backend/scripts/nma_sublocation_source_from_report.py <report.xlsx> [output.xlsx]
"""
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook

from modules.legacy_order.database import get_central_connection
from modules.time_report.excel_export import table_xlsx

_CANONICAL_RE = re.compile(r"^OIN [A-Z]$")
_NORMALIZE_STRIP = re.compile(r"[^A-Z0-9 ]")
_WS = re.compile(r"\s+")


def normalize_name(name):
    """Uppercase, strip punctuation, collapse whitespace, sort tokens --
    a bag-of-words key so word-order differences (BETADINE OINT 20GM vs
    BETADINE 20GM OINT) still match."""
    s = _NORMALIZE_STRIP.sub(" ", (name or "").upper())
    tokens = sorted(_WS.split(s.strip()))
    return " ".join(t for t in tokens if t)


def is_canonical(subloc, product_name):
    """'OIN <single letter>' -- the letter need not match the product name's
    first letter (legacy stores bin by brand/active-ingredient family, e.g.
    BETNOVATE-C sits under 'OIN H', not 'OIN B'). This only rejects the
    genuinely malformed cases (blank, multi-word, non-OIN locations like
    CNT/LOT/J004)."""
    s = (subloc or "").strip().upper()
    return bool(_CANONICAL_RE.match(s))


def read_report_rows(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header = None
    rows = []
    for row in ws.iter_rows(values_only=True):
        if header is None:
            if row and row[0] == "NMC ProductCode":
                header = list(row)
            continue
        if row[0] is None:
            continue
        rows.append(dict(zip(header, row)))
    return rows


def fetch_nma_catalog(cur):
    cur.execute("SELECT ProductCode, ProductName, UnitDescription, SubLocation FROM dbo.Products WHERE StoreName = 'NMA'")
    return [
        {"ProductCode": pc, "ProductName": name, "UnitDescription": unit, "SubLocation": subloc}
        for pc, name, unit, subloc in cur.fetchall()
    ]


def build_name_index(catalog):
    index = {}
    for p in catalog:
        key = normalize_name(p["ProductName"])
        if not key:
            continue
        index.setdefault(key, []).append(p)
    return index


def build(report_path):
    rows = read_report_rows(report_path)

    conn = get_central_connection()
    try:
        cur = conn.cursor()
        nma_catalog = fetch_nma_catalog(cur)
    finally:
        conn.close()
    nma_by_code = {p["ProductCode"]: p for p in nma_catalog}
    nma_name_index = build_name_index(nma_catalog)

    results = []          # (ProductCode, ProductName, SubLocation, UnitDescription, Source)
    seen_nma_codes = set()
    ambiguous_name_matches = []

    for r in rows:
        nmc_name = r.get("NMC ProductName")
        nmc_subloc = r.get("NMC SubLocation")
        if not is_canonical(nmc_subloc, nmc_name):
            continue
        proposed = str(nmc_subloc).strip()

        nma_status = r.get("NMA MappingStatus")
        if nma_status == "MAPPED":
            nma_pc = r.get("NMA ProductCode")
            if nma_pc in seen_nma_codes:
                continue
            p = nma_by_code.get(nma_pc)
            if not p:
                continue
            results.append((p["ProductCode"], p["ProductName"], proposed, p["UnitDescription"], "SUPPLIER_MATCH"))
            seen_nma_codes.add(p["ProductCode"])
            continue

        if nma_status == "MISSING":
            key = normalize_name(nmc_name)
            candidates = nma_name_index.get(key, [])
            if len(candidates) == 1:
                p = candidates[0]
                if p["ProductCode"] in seen_nma_codes:
                    continue
                results.append((p["ProductCode"], p["ProductName"], proposed, p["UnitDescription"], "NAME_MATCH"))
                seen_nma_codes.add(p["ProductCode"])
            elif len(candidates) > 1:
                ambiguous_name_matches.append((r.get("NMC ProductCode"), nmc_name, [c["ProductCode"] for c in candidates]))
        # AMBIGUOUS NMA mapping rows are skipped entirely, same as before.

    return results, ambiguous_name_matches


def export(results, out_path):
    headers = ["ProductCode", "ProductName", "SubLocation", "UnitDescription"]
    xlsx_rows = []
    for pc, name, subloc, unit, source in results:
        # Yellow marks the new normalized-name matches for easy review;
        # green marks the existing SupplierProductMatch-based matches.
        hex_ = "FFEB9C" if source == "NAME_MATCH" else "C6EFCE"
        xlsx_rows.append(([pc, name, subloc, unit or ""], hex_))
    buf = table_xlsx("NMA SubLocation Update Source (report + name-match)", headers, xlsx_rows)
    Path(out_path).write_bytes(buf.getvalue())


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: nma_sublocation_source_from_report.py <report.xlsx> [output.xlsx]")
        sys.exit(1)
    report_path = sys.argv[1]
    out_path = sys.argv[2] if len(sys.argv) > 2 else "nma_sublocation_update_source.xlsx"

    results, ambiguous = build(report_path)

    supplier_matched = [r for r in results if r[4] == "SUPPLIER_MATCH"]
    name_matched = [r for r in results if r[4] == "NAME_MATCH"]

    print(f"Total rows written: {len(results)}")
    print(f"  Via SupplierProductMatch: {len(supplier_matched)}")
    print(f"  Via normalized name match (NEW): {len(name_matched)}")
    print(f"Ambiguous normalized-name matches skipped (>1 NMA candidate): {len(ambiguous)}")
    if name_matched:
        print("\n--- New NAME_MATCH rows (first 30) ---")
        for pc, name, subloc, unit, _ in name_matched[:30]:
            print(f"  {pc} | {name} | -> {subloc}")
    if ambiguous:
        print("\n--- Ambiguous name matches skipped (first 15) ---")
        for nmc_pc, nmc_name, candidates in ambiguous[:15]:
            print(f"  NMC {nmc_pc} ({nmc_name}) -> NMA candidates {candidates}")

    export(results, out_path)
    print(f"\nWritten: {out_path}")
