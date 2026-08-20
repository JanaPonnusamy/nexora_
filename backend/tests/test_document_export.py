"""The workbook a phone hands to an accountant.

`docs/Document_Extraction_Excel_Contract.md` is frozen: "Column order above is
fixed... never reorder or remove a column without a new contract version, since
the whole point of freezing this now is that downstream pharmacy-software
import mappings can be built against it."

So the expectations here are **parsed out of that document** rather than
retyped from the code. A test that mirrors the implementation would pass
whichever way the two drifted apart; this one fails if either side moves.
"""

import io
import re
from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from modules.document_extraction import export_excel

CONTRACT = Path(__file__).resolve().parents[2] / "docs" / "Document_Extraction_Excel_Contract.md"

# Sheet tab names as `build_workbook` writes them, in workbook order. The
# contract fixes the columns; the tab names are asserted here so a rename is a
# deliberate act rather than a silent break for anyone mapping by sheet name.
SHEET_TABS = [
    "Invoice Header",
    "Products",
    "OCR Products",
    "GST Summary",
    "Validation",
    "OCR Metadata",
]


def contract_columns() -> list[list[str]]:
    """Column names per sheet, in order, read out of the frozen contract."""
    text = CONTRACT.read_text(encoding="utf-8")
    sections = re.split(r"^## Sheet \d+ — .*$", text, flags=re.MULTILINE)[1:]
    assert len(sections) == 6, "the contract should describe exactly six sheets"

    sheets = []
    for section in sections:
        columns = []
        for line in section.splitlines():
            # Table body rows look like: | 3 | GST Number | `gst_number` | text | No |
            match = re.match(r"^\|\s*\d+\s*\|\s*([^|]+?)\s*\|", line)
            if match:
                columns.append(match.group(1))
        assert columns, "a sheet section with no column table"
        sheets.append(columns)
    return sheets


@pytest.fixture
def batch():
    """Two invoices in one export — the case the contract exists for, since
    rows from different invoices end up in the same sheets."""
    imports = [
        {
            "import_id": 11,
            "supplier_name": "Sri Balaji Pharma Distributors",
            "gst_number": "33AABCS1429B1ZP",
            "dl_number": "TN/CHE/20B/1234",
            "invoice_number": "SI/26-27/4821",
            "invoice_date": date(2026, 8, 11),
            "invoice_type": "PURCHASE",
            "net_amount": 12480.55,
            "gross_amount": 11890.0,
            "cgst_amount": 295.25,
            "sgst_amount": 295.30,
            "item_count": 2,
            "total_quantity": 100.0,
            "validation_status": "WARNING",
            "source_type": "CAMERA",
            "page_count": 2,
            "ocr_confidence": 86.4,
            "supplier_match_method": "GST",
            "supplier_match_confidence": 97.0,
            "uploaded_at": datetime(2026, 8, 11, 9, 15, 0),
            "reviewed_at": datetime(2026, 8, 11, 9, 41, 0),
            "gst_slab_breakup_json": [
                {"rate": 12, "taxable": 5000.0, "cgst": 300.0, "sgst": 300.0},
                {"rate": 5, "taxable": 6771.1, "cgst": 169.28, "sgst": 169.28},
            ],
            "validation_json": [
                {
                    "rule_code": "MISSING_PRODUCT_CODE",
                    "severity": "WARNING",
                    "item_id": 102,
                    "message": "Product could not be resolved to a ProductCode.",
                },
            ],
        },
        {
            "import_id": 12,
            "supplier_name": "Kumar Medical Agencies",
            "invoice_number": "KMA-9910",
            "invoice_date": date(2026, 8, 12),
            "net_amount": 4300.0,
            "validation_status": "PASSED",
            "source_type": "GALLERY",
            "page_count": 1,
            "uploaded_at": datetime(2026, 8, 12, 10, 0, 0),
            "gst_slab_breakup_json": [],
            "validation_json": [],
        },
    ]
    items = [
        {
            "item_id": 101,
            "import_id": 11,
            "line_number": 1,
            "product_code": "PRD-10041",
            "normalized_product_name": "Dolo 650mg Tablet",
            "pack": "15s",
            "hsn_code": "30049099",
            "batch_number": "DL2411",
            "expiry_date": date(2027, 11, 30),
            "quantity": 40.0,
            "free_quantity": 4.0,
            "purchase_rate": 21.5,
            "mrp": 30.0,
            "gst_percent": 12.0,
            "amount": 860.0,
            "confidence": 0.94,
            "is_excluded": False,
        },
        {
            "item_id": 102,
            "import_id": 11,
            "line_number": 2,
            "ocr_product_name": "AMOXYCLLIN 625 TAB",
            "batch_number": "-",
            "expiry_raw": "13/26",
            "quantity": 60.0,
            "purchase_rate": 92.4,
            "amount": 5544.0,
            "confidence": 0.41,
            "is_excluded": False,
        },
        {
            # A footer line the reviewer removed. It stays in the rows handed to
            # the builder because Sheet 4 findings may still reference it.
            "item_id": 103,
            "import_id": 11,
            "line_number": 3,
            "normalized_product_name": "SUB TOTAL",
            "amount": 6404.0,
            "is_excluded": True,
        },
        {
            "item_id": 201,
            "import_id": 12,
            "line_number": 1,
            "product_code": "PRD-88213",
            "normalized_product_name": "Pantoprazole 40mg Injection",
            "batch_number": "PN0912",
            "expiry_date": date(2027, 2, 28),
            "quantity": 25.0,
            "purchase_rate": 172.0,
            "amount": 4300.0,
            "confidence": 0.88,
            "is_excluded": False,
        },
    ]
    return imports, items


@pytest.fixture
def workbook(batch):
    imports, items = batch
    content = export_excel.build_workbook(
        imports, items, datetime(2026, 8, 17, 8, 30, 0)
    )
    return load_workbook(io.BytesIO(content))


def test_workbook_has_the_six_contract_sheets_in_order(workbook):
    assert workbook.sheetnames == SHEET_TABS


def test_every_sheet_matches_the_frozen_column_list(workbook):
    for tab, expected in zip(SHEET_TABS, contract_columns()):
        actual = [cell.value for cell in workbook[tab][1]]
        assert actual == expected, f"{tab} has drifted from the contract"


def test_every_sheet_leads_with_the_traceability_pair(workbook):
    # The contract's own reason: "Every sheet carries Import ID and Invoice
    # Number as its first two columns specifically so rows from different
    # invoices in the same batch stay traceable after the file leaves this
    # system." Sheet 1 carries the invoice number further along its own row.
    for tab in SHEET_TABS:
        header = [cell.value for cell in workbook[tab][1]]
        if tab == "Products":
            assert header[0] == "No"
            continue
        assert header[0] == "Import ID", tab
        if tab != "Invoice Header":
            assert header[1] == "Invoice Number", tab


def test_one_row_per_invoice_on_the_header_sheet(workbook):
    rows = list(workbook["Invoice Header"].iter_rows(min_row=2, values_only=True))
    assert [row[0] for row in rows] == [11, 12]


def test_products_exclude_removed_lines_but_keep_every_invoice(workbook):
    rows = list(workbook["OCR Products"].iter_rows(min_row=2, values_only=True))

    # Both invoices' lines, minus the one the reviewer removed.
    assert [(row[0], row[2]) for row in rows] == [(11, 1), (11, 2), (12, 1)]
    assert not any(row[4] == "SUB TOTAL" for row in rows)


def test_a_line_falls_back_to_the_raw_ocr_name_and_expiry(workbook):
    rows = list(workbook["OCR Products"].iter_rows(min_row=2, values_only=True))
    unresolved = rows[1]

    assert unresolved[4] == "AMOXYCLLIN 625 TAB"
    # An expiry OCR could not parse still reaches the workbook as what was
    # read, rather than as a blank the accountant cannot chase.
    assert unresolved[8] == "13/26"


def test_products_sheet_matches_the_pharmacy_import_mapping(workbook):
    sheet = workbook["Products"]
    headers = [cell.value for cell in sheet[1]]
    values = dict(zip(headers, next(sheet.iter_rows(min_row=2, values_only=True))))

    assert len(headers) == 79
    assert values["No"] == 1
    assert values["Product Name"] == "Dolo 650mg Tablet"
    assert values["Packing"] == "15s"
    assert values["Batch"] == "DL2411"
    assert values["Qty"] == 40
    assert values["Rate"] == 21.5
    assert values["Goods Value"] == pytest.approx(860)
    assert values["Prod. Code"] == "PRD-10041"
    assert values["Exp.dt"] == "11/27"
    assert values["Mrp"] == 30
    assert values["MRP VALUE"] == pytest.approx(1200)
    assert values["HSN Code"] == "30049099"
    assert values["Rate Changable"] is False


def test_pharmacy_rate_is_recovered_when_invoice_prints_ptr_and_gst_base_only():
    item = {
        "line_number": 1,
        "quantity": 2,
        "ptr": 38.43,
        "gst_percent": 5,
        "amount": 69.13,
    }
    row = dict(zip(export_excel._PRODUCT_COLUMNS, export_excel._product_row(
        item, {"cgst_amount": 48.81, "sgst_amount": 48.81}
    )))

    assert row["Rate"] == pytest.approx(32.919, abs=0.001)
    assert row["Goods Value"] == pytest.approx(65.838, abs=0.001)
    assert row["CGST"] == pytest.approx(1.65)
    assert row["SGST"] == pytest.approx(1.65)
    assert row["Net Rate Amount"] == pytest.approx(69.13)


def test_pharmacy_discount_and_net_values_follow_the_target_system_example():
    item = {
        "line_number": 1,
        "quantity": 5,
        "purchase_rate": 53.3,
        "discount_percent": 12,
        "gst_percent": 5,
        "amount": 246.24,
    }
    row = dict(zip(export_excel._PRODUCT_COLUMNS, export_excel._product_row(
        item, {"cgst_amount": 5.86, "sgst_amount": 5.86}
    )))

    assert row["PDisc.Amount"] == pytest.approx(31.98)
    assert row["SGST"] == pytest.approx(5.86)
    assert row["CGST"] == pytest.approx(5.86)
    assert row["Net Rate Amount"] == pytest.approx(278.22)
    assert row["Net Rate"] == pytest.approx(55.644)


def test_gst_slabs_are_unpacked_one_row_per_slab(workbook):
    rows = list(workbook["GST Summary"].iter_rows(min_row=2, values_only=True))

    assert [(row[0], row[2]) for row in rows] == [(11, 12), (11, 5)]
    # Total GST is derived when the slab does not carry it.
    assert rows[0][7] == pytest.approx(600.0)


def test_findings_resolve_item_ids_to_line_numbers(workbook):
    rows = list(workbook["Validation"].iter_rows(min_row=2, values_only=True))

    assert len(rows) == 1, "an invoice with no findings contributes no rows"
    assert rows[0][0] == 11
    assert rows[0][2] == "MISSING_PRODUCT_CODE"
    # item_id 102 is line 2 — the workbook names the line, not the row id.
    assert rows[0][5] == 2


def test_ocr_metadata_carries_one_row_per_invoice_with_the_export_time(workbook):
    rows = list(workbook["OCR Metadata"].iter_rows(min_row=2, values_only=True))

    assert [row[0] for row in rows] == [11, 12]
    assert all(row[-1] == "2026-08-17 08:30:00" for row in rows)


def test_amounts_are_numeric_cells_not_text(workbook):
    # Binding: "plain numeric Excel cells (not text), so the importing pharmacy
    # software can sum/filter them directly".
    header = workbook["Invoice Header"]
    net_amount = header.cell(row=2, column=22)
    assert isinstance(net_amount.value, (int, float))

    products = workbook["Products"]
    assert isinstance(products.cell(row=2, column=10).value, (int, float))


def test_dates_are_written_in_the_contract_format(workbook):
    assert workbook["Invoice Header"].cell(row=2, column=6).value == "2026-08-11"
    assert workbook["Products"].cell(row=2, column=20).value == "11/27"
    assert workbook["OCR Products"].cell(row=2, column=9).value == "2027-11-30"


def test_header_rows_are_bold_and_frozen(workbook):
    for tab in SHEET_TABS:
        sheet = workbook[tab]
        assert sheet.freeze_panes == "A2", tab
        assert sheet.cell(row=1, column=1).font.bold, tab


def test_the_file_is_a_real_workbook_excel_can_open(batch):
    imports, items = batch
    content = export_excel.build_workbook(imports, items, datetime(2026, 8, 17))

    # xlsx is a zip; the magic bytes are what a phone hands to the share sheet.
    assert content[:2] == b"PK"
    assert len(content) > 4000
    load_workbook(io.BytesIO(content))


def test_csv_export_is_the_products_sheet_with_a_bom(batch):
    imports, items = batch
    content = export_excel.build_products_csv(imports, items)

    assert content.startswith(b"\xef\xbb\xbf"), "Excel needs the BOM to read UTF-8"
    text = content.decode("utf-8-sig")
    lines = text.strip().splitlines()
    assert lines[0].split(",") == contract_columns()[1]
    assert len(lines) == 4, "header plus three non-excluded product lines"
    assert "SUB TOTAL" not in text


def _word(text, center, y):
    width = max(12, len(text) * 7)
    return {
        "text": text,
        "confidence": 0.96,
        "bbox": {"x1": center - width // 2, "y1": y, "x2": center + width // 2, "y2": y + 18},
    }


def _line(line_no, tokens, y):
    words = [_word(text, center, y) for text, center in tokens]
    return {
        "line_no": line_no,
        "text": " ".join(text for text, _center in tokens),
        "confidence": 0.95,
        "bbox": {
            "x1": min(word["bbox"]["x1"] for word in words),
            "y1": y,
            "x2": max(word["bbox"]["x2"] for word in words),
            "y2": y + 18,
        },
        "words": words,
    }


def _source_ocr_json():
    headers = [
        ("S", 15), ("No", 35), ("Description", 130), ("HSN", 260),
        ("Batch", 340), ("Exp", 420), ("Qty", 500), ("Free", 580),
        ("PTR", 660), ("Total", 740),
    ]
    first = [
        ("1", 25), ("ACILOC", 110), ("150", 155), ("30049033", 260),
        ("LD26003", 340), ("06/28", 420), ("2", 500),
        # The Free cell is intentionally blank in the photographed row.
        ("38.43", 660), ("69.13", 740),
    ]
    second = [
        ("2", 25), ("ACITROM", 115), ("2MG", 160), ("30045010", 260),
        ("SSL0213", 340), ("11/27", 420), ("1", 500), ("1", 580),
        ("433.63", 660), ("409.78", 740),
    ]
    return {
        "engine_name": "PaddleOCR",
        "pages": [{
            "page_no": 1,
            "lines": [
                _line(1, headers, 100),
                _line(2, first, 140),
                _line(3, second, 180),
                _line(4, [("SUB", 620), ("TOTAL", 700)], 240),
            ],
        }],
        "average_confidence": 0.95,
    }


def test_source_table_preserves_detected_columns_and_blank_cells(batch):
    imports, items = batch
    imports = [{**imports[0], "ocr_json": _source_ocr_json()}]
    items = [item for item in items if item["import_id"] == 11]

    content = export_excel.build_workbook(imports, items, datetime(2026, 8, 17))
    workbook = load_workbook(io.BytesIO(content))
    sheet = workbook["Source Table"]
    headers = [cell.value for cell in sheet[1]]
    first = list(next(sheet.iter_rows(min_row=2, max_row=2, values_only=True)))
    values = dict(zip(headers, first))

    assert headers == ["S No", "Description", "HSN", "Batch", "Exp", "Qty", "Free", "PTR", "Total"]
    assert values["Description"] == "ACILOC 150"
    assert values["Qty"] == "2"
    assert values["Free"] is None
    assert values["PTR"] == "38.43"
    assert values["Total"] == "69.13"
    assert workbook.active.title == "Source Table"
    assert sheet.freeze_panes == "A2"
    assert sheet.cell(row=2, column=headers.index("Total") + 1).alignment.horizontal == "right"


def test_batch_export_creates_one_source_table_per_invoice(batch):
    imports, items = batch
    imports = [
        {**imports[0], "ocr_json": _source_ocr_json()},
        {**imports[1], "ocr_json": _source_ocr_json()},
    ]

    content = export_excel.build_workbook(imports, items, datetime(2026, 8, 17))
    workbook = load_workbook(io.BytesIO(content))

    assert "Source Table 11" in workbook.sheetnames
    assert "Source Table 12" in workbook.sheetnames
    assert workbook.active.title == "Source Table 11"


def test_malformed_legacy_ocr_does_not_break_contracted_export(batch):
    imports, items = batch
    imports = [{**imports[0], "ocr_json": {"pages": [{"lines": "not-a-list"}]}}]
    items = [item for item in items if item["import_id"] == 11]

    content = export_excel.build_workbook(imports, items, datetime(2026, 8, 17))
    workbook = load_workbook(io.BytesIO(content))

    assert workbook.sheetnames == SHEET_TABS
